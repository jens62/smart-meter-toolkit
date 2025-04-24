#!/usr/bin/env python3

import os
import sys
import re
import logging
import pandas as pd
import argparse
from datetime import datetime
import pytz
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def setup_logging(log_level=logging.INFO):
    """Configure logging with custom formatter and specified log level."""
    logger = logging.getLogger()
    logger.setLevel(log_level)
    
    ch = logging.StreamHandler()
    ch.setLevel(log_level)
    
    formatter = logging.Formatter(
        '%(asctime)-15s %(name)-8s %(lineno)d %(levelname)s: %(message)s'
    )
    ch.setFormatter(formatter)
    
    if not logger.handlers:
        logger.addHandler(ch)
    
    return logger

def load_from_stdin(time_col, value_col, measurement_col, delimiter, divisor, logger):
    """Load and process data from stdin."""
    logger.info("Reading data from stdin...")
    try:
        # Read from stdin
        df = pd.read_csv(sys.stdin, parse_dates=[time_col], delimiter=delimiter)
        
        # Rename columns to standard names
        df = df.rename(columns={
            time_col: '_time',
            value_col: '_value',
            measurement_col: '_measurement'
        })
        
        # Ensure we have the required columns in correct order
        if not all(col in df.columns for col in ['_time', '_value', '_measurement']):
            raise ValueError("Missing required columns in stdin input")
            
        # Reorder columns consistently
        df = df[['_time', '_value', '_measurement']]
            
        if divisor != 1:
            logger.info(f"Dividing _value column by {divisor}")
            df['_value'] = df['_value'] / divisor
            
        return df
        
    except Exception as e:
        logger.error(f"Error reading from stdin: {str(e)}")
        raise

def load_and_process_data(input_path, file_pattern, timezone, column_mapping, divisor, logger, delimiter=','):
    """Load and process data from file(s)."""
    target_tz = pytz.timezone(timezone)
    
    if os.path.isfile(input_path):
        logger.info(f"Processing single file: {input_path}")
        df = pd.read_csv(input_path, parse_dates=[column_mapping['_time']], delimiter=delimiter)
        df = rename_columns(df, column_mapping)
        df = df[['_time', '_value', '_measurement']]
    else:
        logger.info(f"Processing files in directory: {input_path} with pattern: {file_pattern}")
        data_frames = []
        pattern = re.compile(file_pattern)
        
        for file in os.listdir(input_path):
            if pattern.search(file):
                file_path = os.path.join(input_path, file)
                try:
                    df = pd.read_csv(file_path, parse_dates=[column_mapping['_time']], delimiter=delimiter)
                    df = rename_columns(df, column_mapping)
                    df = df[['_time', '_value', '_measurement']]
                    data_frames.append(df)
                except Exception as e:
                    logger.error(f"Error processing file {file}: {str(e)}")
                    continue
        
        if not data_frames:
            raise ValueError(f"No files matching pattern '{file_pattern}' found.")
        
        df = pd.concat(data_frames, ignore_index=True)
    
    # Process data
    try:
        # Ensure timestamps are timezone-aware (UTC)
        if df["_time"].dt.tz is None:
            logger.info("Localizing naive timestamps to UTC")
            df["_time"] = df["_time"].dt.tz_localize('UTC')
        
        # Convert to target timezone
        logger.info(f"Converting timestamps to target timezone: {timezone}")
        df["_time"] = df["_time"].dt.tz_convert(target_tz)
        
        # Remove timezone info for Excel compatibility
        df['_time'] = df['_time'].dt.tz_localize(None)
    
    except Exception as e:
        logger.error(f"Error processing timestamps: {str(e)}")
        raise
    
    if divisor != 1:
        logger.info(f"Dividing _value column by {divisor}")
        df['_value'] = df['_value'] / divisor
    
    return df.drop_duplicates()

def rename_columns(df, column_mapping):
    """Rename columns according to mapping."""
    return df.rename(columns={v: k for k, v in column_mapping.items()})

def generate_json_output(df, output_dir, measurement_name, unit, timezone):
    """Generate JSON output with consumption data."""
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.json"
    
    # Make a copy with timezone-aware timestamps for JSON output
    tz = pytz.timezone(timezone)
    df_json = df.copy()
    df_json['_time'] = pd.to_datetime(df_json['_time']).dt.tz_localize(tz)
    
    output_data = {"consumption": []}
    
    # Group by year and month in target timezone
    df_json["_year"] = df_json["_time"].dt.strftime("%Y")
    df_json["_month"] = df_json["_time"].dt.strftime("%m")
    
    for (year, month), month_df in df_json.groupby(["_year", "_month"]):
        max_val = month_df["_value"].max()
        min_val = month_df["_value"].min()
        consumption = max_val - min_val
        
        meter_values = []
        for _, row in month_df.iterrows():
            meter_values.append({
                "time": row["_time"].isoformat(),
                "value": round(row["_value"], 4),
                "meter_id": row["_measurement"]
            })
        
        output_data["consumption"].append({
            "year": year,
            "month": month,
            "consumption_month": round(consumption, 4),
            "meter_values": meter_values
        })
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        json.dump(output_data, f, indent=2)
    
    return output_filename

def generate_excel_output(df, output_dir, measurement_name, unit, timezone):
    """Generate Excel output with professional German formatting."""
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.xlsx"
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Apply timezone to timestamps for correct grouping
    tz = pytz.timezone(timezone)
    df_excel = df.copy()
    df_excel['_time'] = pd.to_datetime(df_excel['_time']).dt.tz_convert(timezone)

    
    # ===== STYLE DEFINITIONS =====
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    header_style = {
        'font': header_font,
        'fill': header_fill,
        'border': thin_border
    }
    
    # ===== VERBRAUCH SHEET =====
    ws_consumption = wb.create_sheet(title="Verbrauch")
    ws_consumption.append(["Monat", f"Verbrauch [{unit}]"])
    
    # Apply header style
    for cell in ws_consumption[1]:
        for attr, value in header_style.items():
            setattr(cell, attr, value)
    
    # ===== MONTHLY SHEETS =====
    df_excel["_year_month"] = df_excel["_time"].dt.strftime("%Y_%m")
    for year_month, month_df in df_excel.groupby("_year_month"):
        # Create worksheet with German headers
        ws = wb.create_sheet(title=year_month)
        ws.append(["Zeit der Messung", f"Zählerstand [{unit}]", "Zähleridentifikation"])
        
        # Get column indices once
        time_idx = month_df.columns.get_loc('_time')
        value_idx = month_df.columns.get_loc('_value')
        meas_idx = month_df.columns.get_loc('_measurement')
        
        # Add data rows using column indices
        for row in month_df.itertuples(index=False):
            ws.append([
                row[time_idx].strftime("%d.%m.%Y %H:%M:%S"),
                row[value_idx],
                row[meas_idx]
            ])
        
        # Apply German number formatting
        apply_german_number_format(ws, value_col_idx=2)
        
        # Apply zebra formatting with borders
        apply_zebra_formatting(ws, header_style)
        
        # Set dynamic column widths
        set_dynamic_column_widths(ws, month_df)
        
        # Add consumption formula
        formula = f"=MAX('{year_month}'!B:B)-MIN('{year_month}'!B:B)"
        ws_consumption.append([year_month, formula])
    
    # ===== FINALIZE VERBRAUCH SHEET =====
    # Apply formatting to consumption sheet
    apply_german_number_format(ws_consumption, value_col_idx=2)
    apply_zebra_formatting(ws_consumption, header_style)
    set_dynamic_column_widths(ws_consumption, df, is_consumption=True)
    
    # Save workbook
    wb.save(os.path.join(output_dir, output_filename))
    return output_filename

def generate_csv_output(df, output_dir, measurement_name):
    """Generate CSV output file."""
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.csv"
    
    df.to_csv(os.path.join(output_dir, output_filename), index=False)
    return output_filename

def generate_output_files(df, output_dir, logger, timezone, measurement_name=None, unit="kWh", out_formats=["excel"]):
    """Generate output files in specified formats."""
    # If measurement_name not provided and dataframe is empty, use "unknown"
    if measurement_name is None:
        if df.empty:
            measurement_name = "unknown"
        else:
            measurement_name = df["_measurement"].iloc[0] if "_measurement" in df.columns else "unknown"
    
    output_files = []
    
    if "csv" in out_formats:
        try:
            csv_file = generate_csv_output(df, output_dir, measurement_name)
            output_files.append(csv_file)
            logger.info(f"Created CSV file: {csv_file}")
        except Exception as e:
            logger.error(f"Failed to generate CSV output: {str(e)}")
    
    if "json" in out_formats:
        try:
            json_file = generate_json_output(df, output_dir, measurement_name, unit, timezone)
            output_files.append(json_file)
            logger.info(f"Created JSON file: {json_file}")
        except Exception as e:
            logger.error(f"Failed to generate JSON output: {str(e)}")
    
    if "excel" in out_formats:
        try:
            excel_file = generate_excel_output(df, output_dir, measurement_name, unit, timezone)
            output_files.append(excel_file)
            logger.info(f"Created Excel file: {excel_file}")
        except Exception as e:
            logger.error(f"Failed to generate Excel output: {str(e)}")
    
    return output_files

def apply_german_number_format(ws, value_col_idx):
    """Apply German number formatting to value column"""
    german_number_format = '#,##0.0000;[Red]-#,##0.0000'
    for row in ws.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.value = round(cell.value, 4)  # Ensure 4 decimal places
                cell.number_format = german_number_format

def apply_zebra_formatting(ws, header_style):
    """Apply zebra striping with borders to all cells"""
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            # Header row
            if cell.row == 1:
                for attr, value in header_style.items():
                    setattr(cell, attr, value)
            # Data rows
            else:
                fill = light_fill if cell.row % 2 == 0 else white_fill
                cell.fill = fill
                cell.border = header_style['border']

def set_dynamic_column_widths(ws, df, is_consumption=False):
    """Set optimized column widths including header text"""
    if is_consumption:
        # Verbrauch sheet columns
        header_lengths = {
            'A': len("Monat"),
            'B': len(ws['B1'].value)  # "Verbrauch [kWh]"
        }
        content_lengths = {
            'A': max(len(str(v.value)) for v in ws['A']),
            'B': max(len(f"{v.value:,.4f}") if isinstance(v.value, (int, float)) else len(str(v.value)) 
                   for v in ws['B'])
        }
        
        # Use whichever is longer - header or content
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 15)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)
        
    else:
        # Monthly sheet columns
        headers = ["Zeit der Messung", f"Zählerstand [kWh]", "Zähleridentifikation"]
        header_lengths = {
            'A': len(headers[0]),
            'B': len(headers[1]),
            'C': len(headers[2])
        }
        
        content_lengths = {
            'A': max(len(row[df.columns.get_loc('_time')].strftime("%d.%m.%Y %H:%M:%S")) for row in df.itertuples(index=False)),
            'B': max(len(f"{v:,.4f}") if isinstance(v, (int, float)) else len(str(v)) 
                   for v in df["_value"]),
            'C': max(len(str(v)) for v in df["_measurement"])
        }
        
        # Apply widths with header consideration
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 22)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)
        ws.column_dimensions['C'].width = min(max(header_lengths['C'], content_lengths['C']) + 2, 35)

def main():
    examples = """
Examples:
  # Process with default unit (kWh) and output formats
  meter_csv2excel.py --file data.csv --time-col capture_time --value-col long64_value --measurement-col logical_name --delimiter ';'
  
  # Process with custom unit (MWh) and specific output formats
  meter_csv2excel.py --file data.csv --unit "MWh" --divisor 1000 --out-format excel json

  # Process from stdin with JSON output only
  cat data.csv | meter_csv2excel.py --stdin --time-col capture_time --value-col long64_value --measurement-col logical_name --delimiter ';' --out-format json
"""
    
    parser = argparse.ArgumentParser(
        description="Merge CSV files, convert timezones, and generate reports in multiple formats.",
        epilog=examples,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Input options - make mutually exclusive with stdin
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument("--folder", help="Path to folder with CSV files")
    input_group.add_argument("--file", help="Path to single CSV file")
    input_group.add_argument("--stdin", action="store_true", 
                           help="Read input from stdin instead of file")
    
    # Processing parameters
    parser.add_argument("--timezone", default="Europe/Berlin",
                       help="Timezone for conversion (default: Europe/Berlin)")
    parser.add_argument("--delimiter", default=",",
                       help="CSV delimiter character (default: ',')")
    parser.add_argument("--pattern", default=r".*_reduced\.csv$",
                       help="File pattern regex (default: '.*_reduced.csv$')")
    parser.add_argument("--output", default=".",
                       help="Output directory (default: current)")
    
    # Column mappings
    parser.add_argument("--time-col", default="_time",
                       help="Time column name (default: '_time')")
    parser.add_argument("--value-col", default="_value",
                       help="Value column name (default: '_value')")
    parser.add_argument("--measurement-col", default="_measurement",
                       help="Measurement column name (default: '_measurement')")
    
    # Processing options
    parser.add_argument("--divisor", type=float, default=1.0,
                       help="Divide values by this number (default: 1)")
    parser.add_argument("--log-level", default="INFO",
                       choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
                       help="Logging level (default: INFO)")
    
    # Output options
    parser.add_argument("--unit", default="kWh",
                       help="Measurement unit for display (default: kWh)")
    parser.add_argument("--out-format", nargs="+", default=["excel"],
                       choices=["excel", "csv", "json", "xml"],
                       help="Output formats (default: excel)")
    
    args = parser.parse_args()

    logger = setup_logging(getattr(logging, args.log_level))
    
    try:
        if args.stdin:
            df = load_from_stdin(
                time_col=args.time_col,
                value_col=args.value_col,
                measurement_col=args.measurement_col,
                delimiter=args.delimiter,
                divisor=args.divisor,
                logger=logger
            )
        else:
            column_mapping = {
                '_time': args.time_col,
                '_value': args.value_col,
                '_measurement': args.measurement_col
            }
            
            df = load_and_process_data(
                input_path=args.file or args.folder,
                file_pattern=args.pattern,
                timezone=args.timezone,
                column_mapping=column_mapping,
                divisor=args.divisor,
                logger=logger,
                delimiter=args.delimiter
            )
        
        output_files = generate_output_files(
            df, 
            args.output, 
            logger, 
            unit=args.unit,
            out_formats=args.out_format,
            timezone=args.timezone
        )
        logger.info(f"Successfully created files: {', '.join(output_files)}")
    
    except Exception as e:
        logger.error(f"Processing failed: {str(e)}", exc_info=logger.level == logging.DEBUG)
        exit(1)

if __name__ == "__main__":
    main()
