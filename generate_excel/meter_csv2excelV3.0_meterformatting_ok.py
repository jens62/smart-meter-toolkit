#!/usr/bin/env python3

import os
import sys
import re
import logging
import pandas as pd
import argparse
from datetime import datetime
import pytz
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def setup_logging(log_level=logging.INFO):
    """Configure logging with custom formatter and specified log level."""
    logger = logging.getLogger()
    logger.setLevel(log_level)
    
    ch = logging.StreamHandler()
    ch.setLevel(log_level)
    
    formatter = logging.Formatter(
        '%(asctime)-15s %(name)-8s %(lineno)d %(levelname)s: %(message)s'
    )
    ch.setFormatter(formatter)
    
    if not logger.handlers:
        logger.addHandler(ch)
    
    return logger

# Extract part between dots and format it, if length is 14 chars, according to DIN 43863-5
# Zeichen 1: Spartenkennung
# Zeichen 2-4: Herstellerkennzeichnung
# Zeichen 5-6: Fabrikationsblock
# Zeichen 7-14: Fabrikationsnummer
def format_measurement(x):
    parts = x.split(".")
    if len(parts) > 1 and len(parts[1]) == 14:
        return re.sub(r'^(.{1})(.{3})(.{2})(.{4})(.{4})$', 
                     r'\1_\2_\3_\4_\5', 
                     parts[1]).upper()
    return x.rsplit(".", 1)[0]

def convert_to_timezone(df, timezone, logger=None):
    """Convert dataframe timestamps to specified timezone."""
    if logger and logger.level <= logging.DEBUG:
        logger.debug(f"Converting timestamps to timezone: {timezone}")
    
    # Ensure timestamps are timezone-aware (UTC)
    if df["_time"].dt.tz is None:
        df["_time"] = df["_time"].dt.tz_localize('UTC')
    
    # Convert to target timezone
    tz = pytz.timezone(timezone)
    df["_time"] = df["_time"].dt.tz_convert(tz)
    return df

def load_from_stdin(time_col, value_col, measurement_col, delimiter, divisor, logger=None):
    """Load and process data from stdin."""
    if logger:
        logger.info("Reading data from stdin...")
    
    try:
        df = pd.read_csv(sys.stdin, parse_dates=[time_col], delimiter=delimiter)
        df = df.rename(columns={
            time_col: '_time',
            value_col: '_value',
            measurement_col: '_measurement'
        })

        # df['_measurement'] = df['_measurement'].str.split(".").str[1]  # Extract part between dots
        df['_measurement'] = df['_measurement'].apply(format_measurement)

        # 1. Sort by _time in descending order (youngest first)
        df = df.sort_values('_time', ascending=False)

        # 2. Remove duplicate rows (keeping the first occurrence - which will be the youngest due to sorting)
        df = df.drop_duplicates()      
       
        # Ensure required columns exist
        if not all(col in df.columns for col in ['_time', '_value', '_measurement']):
            raise ValueError("Missing required columns in input data")
            
        # Reorder columns
        df = df[['_time', '_value', '_measurement']]
   
        
        if divisor != 1:
            df['_value'] = df['_value'] / divisor
            
        return df
    except Exception as e:
        if logger:
            logger.error(f"Error reading from stdin: {str(e)}")
        raise

def load_and_process_data(input_path, file_pattern, timezone, column_mapping, divisor, logger=None, delimiter=','):
    """Load and process data from file(s)."""
    if logger:
        logger.info(f"Processing data with timezone: {timezone}")
    
    if os.path.isfile(input_path):
        df = pd.read_csv(input_path, parse_dates=[column_mapping['_time']], delimiter=delimiter)
    else:
        data_frames = []
        pattern = re.compile(file_pattern)
        
        for file in os.listdir(input_path):
            if pattern.search(file):
                try:
                    df = pd.read_csv(os.path.join(input_path, file), 
                                   parse_dates=[column_mapping['_time']], 
                                   delimiter=delimiter)
                    data_frames.append(df)
                except Exception as e:
                    if logger:
                        logger.error(f"Error processing file {file}: {str(e)}")
                    continue
        
        if not data_frames:
            raise ValueError(f"No files matching pattern '{file_pattern}' found")
        df = pd.concat(data_frames, ignore_index=True)
    
    # Rename columns
    df = rename_columns(df, column_mapping)
    df = df[['_time', '_value', '_measurement']]

    # df['_measurement'] = df['_measurement'].str.split(".").str[1]  # Extract part between dots
    df['_measurement'] = df['_measurement'].apply(format_measurement)

    # 1. Sort by _time in descending order (youngest first)
    df = df.sort_values('_time', ascending=False)

    # 2. Remove duplicate rows (keeping the first occurrence - which will be the youngest due to sorting)
    df = df.drop_duplicates()        
    
    # Process data
    try:
        df = convert_to_timezone(df, timezone, logger)
        df['_time'] = df['_time'].dt.tz_localize(None)  # Remove tz for processing
        
        if divisor != 1:
            df['_value'] = df['_value'] / divisor
            
        return df.drop_duplicates()
    except Exception as e:
        if logger:
            logger.error(f"Error processing data: {str(e)}")
        raise

def rename_columns(df, column_mapping):
    """Rename columns according to mapping."""
    return df.rename(columns={v: k for k, v in column_mapping.items()})

def apply_german_number_format(ws, value_col_idx):
    """Apply German number formatting to value column"""
    german_number_format = '#,##0.0000;[Red]-#,##0.0000'
    for row in ws.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.value = round(cell.value, 4)  # Ensure 4 decimal places
                cell.number_format = german_number_format

def apply_zebra_formatting(ws, header_style):
    """Apply zebra striping with borders to all cells"""
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            # Header row
            if cell.row == 1:
                for attr, value in header_style.items():
                    setattr(cell, attr, value)
            # Data rows
            else:
                fill = light_fill if cell.row % 2 == 0 else white_fill
                cell.fill = fill
                cell.border = header_style['border']

def set_dynamic_column_widths(ws, df, is_consumption=False):
    """Set optimized column widths including header text"""
    if is_consumption:
        # Verbrauch sheet columns
        header_lengths = {
            'A': len("Monat"),
            'B': len(ws['B1'].value)  # "Verbrauch [kWh]"
        }
        content_lengths = {
            'A': max(len(str(v.value)) for v in ws['A']),
            'B': max(len(f"{v.value:,.4f}") if isinstance(v.value, (int, float)) else len(str(v.value)) 
                   for v in ws['B'])
        }
        
        # Use whichever is longer - header or content
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 15)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)
        
    else:
        # Monthly sheet columns
        headers = ["Zeit der Messung", f"Zählerstand [kWh]", "Zähleridentifikation"]
        header_lengths = {
            'A': len(headers[0]),
            'B': len(headers[1]),
            'C': len(headers[2])
        }
        
        content_lengths = {
            'A': max(len(row[df.columns.get_loc('_time')].strftime("%d.%m.%Y %H:%M:%S")) for row in df.itertuples(index=False)),
            'B': max(len(f"{v:,.4f}") if isinstance(v, (int, float)) else len(str(v)) 
                   for v in df["_value"]),
            'C': max(len(str(v)) for v in df["_measurement"])
        }
        
        # Apply widths with header consideration
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 22)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)
        ws.column_dimensions['C'].width = min(max(header_lengths['C'], content_lengths['C']) + 2, 35)

def generate_excel_output(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate Excel output with professional German formatting."""
    try:
        if logger:
            logger.info("Generating Excel output...")
        
        time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
        output_filename = f"{measurement_name}_{time_range}.xlsx"
        
        # Make a copy and convert to timezone for grouping
        df_excel = df.copy()
        df_excel = convert_to_timezone(df_excel, timezone, logger)
        
        # Remove timezone for Excel compatibility
        df_excel['_time'] = df_excel['_time'].dt.tz_localize(None)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # ===== STYLE DEFINITIONS =====
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        header_style = {
            'font': header_font,
            'fill': header_fill,
            'border': thin_border
        }
        
        # ===== VERBRAUCH SHEET =====
        ws_consumption = wb.create_sheet(title="Verbrauch")
        ws_consumption.append(["Monat", f"Verbrauch [{unit}]"])
        
        # Apply header style
        for cell in ws_consumption[1]:
            for attr, value in header_style.items():
                setattr(cell, attr, value)
        
        # ===== MONTHLY SHEETS =====
        df_excel["_year_month"] = df_excel["_time"].dt.strftime("%Y_%m")
        for year_month, month_df in df_excel.groupby("_year_month"):
            # Create worksheet with German headers
            ws = wb.create_sheet(title=year_month)
            ws.append(["Zeit der Messung", f"Zählerstand [{unit}]", "Zähleridentifikation"])
            
            # Get column indices once
            time_idx = month_df.columns.get_loc('_time')
            value_idx = month_df.columns.get_loc('_value')
            meas_idx = month_df.columns.get_loc('_measurement')
            
            # Add data rows using column indices
            for row in month_df.itertuples(index=False):
                ws.append([
                    row[time_idx].strftime("%d.%m.%Y %H:%M:%S"),
                    row[value_idx],
                    row[meas_idx]
                ])
            
            # Apply German number formatting
            apply_german_number_format(ws, value_col_idx=2)
            
            # Apply zebra formatting with borders
            apply_zebra_formatting(ws, header_style)
            
            # Set dynamic column widths
            set_dynamic_column_widths(ws, month_df)
            
            # Add consumption formula
            formula = f"=MAX('{year_month}'!B:B)-MIN('{year_month}'!B:B)"
            ws_consumption.append([year_month, formula])
        
        # ===== FINALIZE VERBRAUCH SHEET =====
        # Apply formatting to consumption sheet
        apply_german_number_format(ws_consumption, value_col_idx=2)
        apply_zebra_formatting(ws_consumption, header_style)
        set_dynamic_column_widths(ws_consumption, df_excel, is_consumption=True)
        
        # Save workbook
        wb.save(os.path.join(output_dir, output_filename))
        return output_filename
        
    except Exception as e:
        if logger:
            logger.error(f"Error generating Excel output: {str(e)}")
        raise

def generate_csv_output(df, measurement_name, unit, timezone, logger=None):
    """Generate CSV formatted output with period and consumption data."""
    try:
        if logger:
            logger.debug("Generating CSV output with period data")
        
        # Make timezone-aware copy
        df_csv = df.copy()
        df_csv = convert_to_timezone(df_csv, timezone, logger)
        
        # Add year and month columns
        df_csv["year"] = df_csv["_time"].dt.strftime("%Y")
        df_csv["month"] = df_csv["_time"].dt.strftime("%m")
        
        # Calculate min and max values per period
        grouped = df_csv.groupby(["year", "month", "_measurement"])
        min_values = grouped["_value"].min().reset_index()
        max_values = grouped["_value"].max().reset_index()
        
        # Calculate consumption per period
        consumption = max_values.copy()
        consumption["period_consumption"] = max_values["_value"] - min_values["_value"]
        
        # Merge with original data
        result = pd.merge(df_csv, consumption[["year", "month", "_measurement", "period_consumption"]],
                         on=["year", "month", "_measurement"])
        
        # Rename columns for output
        result = result.rename(columns={
            "_time": "time",
            "_value": "reading",
            "_measurement": "meter_id"
        })
        
        # Add unit information
        result["reading_unit"] = unit
        result["consumption_unit"] = unit
        
        # Select and order columns
        output_columns = [
            "time", "reading", "reading_unit", 
            "meter_id", "year", "month", 
            "period_consumption", "consumption_unit"
        ]
        
        return result[output_columns].to_csv(index=False)
    except Exception as e:
        if logger:
            logger.error(f"Error generating CSV: {str(e)}")
        raise

def generate_json_output(df, measurement_name, unit, timezone, logger=None):
    """Generate JSON formatted output."""
    try:
        if logger:
            logger.debug("Generating JSON output")
        
        # Make timezone-aware copy
        df_json = df.copy()
        df_json = convert_to_timezone(df_json, timezone, logger)
        
        output_data = {"consumption": []}
        
        # Group by year and month
        df_json["_year"] = df_json["_time"].dt.strftime("%Y")
        df_json["_month"] = df_json["_time"].dt.strftime("%m")
        
        for (year, month), month_df in df_json.groupby(["_year", "_month"]):
            max_val = month_df["_value"].max()
            min_val = month_df["_value"].min()
            consumption = max_val - min_val
            
            meter_values = []
            for _, row in month_df.iterrows():
                meter_values.append({
                    "time": row["_time"].isoformat(),
                    "reading": {
                        "value": round(row["_value"], 4),
                        "unit": unit
                    },
                    "meter_id": row["_measurement"]
                })
            
            output_data["consumption"].append({
                "year": year,
                "month": month,
                "consumption_month": {
                    "value": round(consumption, 4),
                    "unit": unit
                },
                "meter_values": meter_values
            })
        
        return json.dumps(output_data, indent=2)
    except Exception as e:
        if logger:
            logger.error(f"Error generating JSON: {str(e)}")
        raise

def generate_xml_output(df, measurement_name, unit, timezone, logger=None):
    """Generate XML formatted output."""
    try:
        if logger:
            logger.debug("Generating XML output")
        
        # Make timezone-aware copy
        df_xml = df.copy()
        df_xml = convert_to_timezone(df_xml, timezone, logger)
        
        # Create root element
        root = ET.Element("consumption_data")
        
        # Group by year and month
        df_xml["_year"] = df_xml["_time"].dt.strftime("%Y")
        df_xml["_month"] = df_xml["_time"].dt.strftime("%m")
        
        for (year, month), month_df in df_xml.groupby(["_year", "_month"]):
            max_val = month_df["_value"].max()
            min_val = month_df["_value"].min()
            consumption = max_val - min_val
            
            period = ET.SubElement(root, "period")
            ET.SubElement(period, "year").text = year
            ET.SubElement(period, "month").text = month
            
            consumption_month = ET.SubElement(period, "consumption_month")
            ET.SubElement(consumption_month, "value").text = str(round(consumption, 4))
            ET.SubElement(consumption_month, "unit").text = unit
            
            meter_values = ET.SubElement(period, "meter_values")
            for _, row in month_df.iterrows():
                item = ET.SubElement(meter_values, "item")
                ET.SubElement(item, "time").text = row["_time"].isoformat()
                
                reading = ET.SubElement(item, "reading")
                ET.SubElement(reading, "value").text = str(round(row["_value"], 4))
                ET.SubElement(reading, "unit").text = unit
                
                ET.SubElement(item, "meter_id").text = row["_measurement"]
        
        # Create pretty printed XML
        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")
    except Exception as e:
        if logger:
            logger.error(f"Error generating XML: {str(e)}")
        raise

def generate_output_files(df, output_dir, logger, measurement_name=None, unit="kWh", out_formats=None, timezone=None):
    """Generate output files in specified formats."""
    if out_formats is None:
        out_formats = ["excel"]
    
    if "none" in out_formats:
        return []
    
    if measurement_name is None:
        measurement_name = df["_measurement"].iloc[0] if "_measurement" in df.columns else "unknown"
    
    output_files = []
    
    if "excel" in out_formats:
        try:
            excel_file = generate_excel_output(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(excel_file)
            logger.info(f"Created Excel file: {excel_file}")
        except Exception as e:
            logger.error(f"Failed to generate Excel output: {str(e)}")
    
    if "json" in out_formats:
        try:
            json_file = generate_json_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(json_file)
            logger.info(f"Created JSON file: {json_file}")
        except Exception as e:
            logger.error(f"Failed to generate JSON output: {str(e)}")
    
    if "xml" in out_formats:
        try:
            xml_file = generate_xml_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(xml_file)
            logger.info(f"Created XML file: {xml_file}")
        except Exception as e:
            logger.error(f"Failed to generate XML output: {str(e)}")
    
    if "csv" in out_formats:
        try:
            csv_file = generate_csv_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(csv_file)
            logger.info(f"Created CSV file: {csv_file}")
        except Exception as e:
            logger.error(f"Failed to generate CSV output: {str(e)}")
    
    return output_files

def generate_csv_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate CSV output to file."""
    csv_content = generate_csv_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.csv"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(csv_content)
    
    return output_filename

def generate_json_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate JSON output to file."""
    json_content = generate_json_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.json"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(json_content)
    
    return output_filename

def generate_xml_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate XML output to file."""
    xml_content = generate_xml_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.xml"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(xml_content)
    
    return output_filename

def main():
    examples = """
Examples:
  # Output XML to stdout only
  cat data.csv | meter_csv2excel.py --stdin --out-format none --stdout-format xml
  
  # Output to both files (Excel and XML) and JSON to stdout
  meter_csv2excel.py --file data.csv --out-format excel xml --stdout-format json
  
  # Output CSV to stdout
  meter_csv2excel.py --file data.csv --stdout-format csv
  
  # Output CSV to file
  meter_csv2excel.py --file data.csv --out-format csv
  
  # Default behavior (Excel file only)
  meter_csv2excel.py --file data.csv
"""
    
    parser = argparse.ArgumentParser(
        description="Process meter data and generate reports",
        epilog=examples,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Input options
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument("--folder", help="Path to folder with CSV files")
    input_group.add_argument("--file", help="Path to single CSV file")
    input_group.add_argument("--stdin", action="store_true", help="Read from stdin")
    
    # Processing parameters
    parser.add_argument("--timezone", default="Europe/Berlin", help="Timezone for conversion")
    parser.add_argument("--delimiter", default=",", help="CSV delimiter character")
    parser.add_argument("--pattern", default=r".*\.csv$", help="File pattern regex")
    parser.add_argument("--output", default=".", help="Output directory")
    
    # Column mappings
    parser.add_argument("--time-col", default="_time", help="Time column name")
    parser.add_argument("--value-col", default="_value", help="Value column name")
    parser.add_argument("--measurement-col", default="_measurement", help="Measurement column name")
    
    # Processing options
    parser.add_argument("--divisor", type=float, default=1.0, help="Divide values by this number")
    parser.add_argument("--unit", default="kWh", help="Measurement unit for display")
    
    # Output options
    parser.add_argument("--out-format", nargs="+", default=["excel"],
                       choices=["none", "excel", "csv", "json", "xml"],
                       help="Output file formats")
    parser.add_argument("--stdout-format", default="none",
                       choices=["none", "json", "csv", "xml"],
                       help="Output format for stdout")
    
    # Logging
    parser.add_argument("--log-level", default="INFO",
                       choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
                       help="Logging level")
    
    args = parser.parse_args()
    logger = setup_logging(getattr(logging, args.log_level))
    
    try:
        # Load data
        if args.stdin:
            df = load_from_stdin(
                args.time_col, args.value_col, args.measurement_col,
                args.delimiter, args.divisor, logger
            )
        else:
            column_mapping = {
                '_time': args.time_col,
                '_value': args.value_col,
                '_measurement': args.measurement_col
            }
            df = load_and_process_data(
                args.file or args.folder, args.pattern, args.timezone,
                column_mapping, args.divisor, logger, args.delimiter
            )
        
        # Handle stdout output
        if args.stdout_format != "none":
            if args.stdout_format == "json":
                try:
                    json_output = generate_json_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(json_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout JSON: {str(e)}")
                    sys.exit(1)
            elif args.stdout_format == "xml":
                try:
                    xml_output = generate_xml_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(xml_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout XML: {str(e)}")
                    sys.exit(1)
            elif args.stdout_format == "csv":
                try:
                    csv_output = generate_csv_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(csv_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout CSV: {str(e)}")
                    sys.exit(1)
        
        # Handle file outputs
        output_files = generate_output_files(
            df, args.output, logger, None, args.unit, args.out_format, args.timezone
        )
        
        if not output_files and "none" not in args.out_format:
            logger.error("No output files were created")
            sys.exit(1)
            
    except Exception as e:
        logger.error(f"Processing failed: {str(e)}", exc_info=logger.level <= logging.DEBUG)
        sys.exit(1)

if __name__ == "__main__":
    main()