#!/usr/bin/env python3
import argparse
import sys
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pytz
from io import StringIO
import os
import json

# Configure logging to stderr
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', stream=sys.stderr)
logger = logging.getLogger(__name__)

def parse_delta(delta_str):
    """Parse delta string into timedelta."""
    units = {
        's': 'seconds',
        'm': 'minutes',
        'h': 'hours',
        'd': 'days',
        'w': 'weeks',
        'M': 'days'  # Approximate month as 30 days
    }
    
    try:
        value = int(delta_str[:-1])
        unit = delta_str[-1]
        if unit not in units:
            raise ValueError(f"Invalid unit '{unit}'. Use s, m, h, d, w, or M.")
        if unit == 'M':
            return timedelta(days=value * 30)
        return timedelta(**{units[unit]: value})
    except (ValueError, IndexError) as e:
        raise argparse.ArgumentTypeError(f"Invalid delta format '{delta_str}'. Expected format like '2h', '30m', etc.") from e

def detect_gaps(df, time_col, delta, timezone):
    """Detect gaps in timestamp column."""
    # Convert to timezone
    tz = pytz.timezone(timezone)
    df['local_time'] = df[time_col].dt.tz_convert(tz)
    
    # Sort by time (ascending)
    df = df.sort_values(time_col)
    
    # Calculate time differences
    df['time_diff'] = df[time_col].diff()
    
    # Identify gaps
    gaps = df[df['time_diff'] > delta]
    
    # Prepare gap information
    gap_list = []
    for _, row in gaps.iterrows():
        gap_start = row[time_col] - row['time_diff']
        gap_end = row[time_col]
        gap_list.append({
            'from': gap_start.astimezone(tz).isoformat(),
            'to': gap_end.astimezone(tz).isoformat()
        })
    
    return gap_list

def style_excel_sheet(ws, gaps):
    """Apply styling to the Excel worksheet."""
    # Zebra styling
    light_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    dark_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Apply header style to all cells in first row
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Apply data styles
    for row_idx, _ in enumerate(gaps, start=2):
        fill = light_fill if row_idx % 2 == 0 else dark_fill
        for col in range(1, 3):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def generate_output_filename(args, from_time, to_time):
    """Generate output filename based on parameters."""
    filename = f"{args.prefix}gaps_between_{from_time}_and_{to_time}"
    if args.format == 'excel':
        return f"{filename}.xlsx"
    elif args.format == 'json':
        return f"{filename}.json"
    elif args.format == 'text':
        return f"{filename}.txt"
    return filename

def write_output(args, output_data, from_time, to_time):
    """Handle output based on format and destination."""
    output_filename = None
    
    if args.output_dest in ['file', 'both']:
        output_filename = os.path.join(args.out, generate_output_filename(args, from_time, to_time))
        os.makedirs(args.out, exist_ok=True)
        
        if args.format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "gaps"
            ws.append(['Start', 'End'])
            for gap in output_data['gaps']:
                ws.append([gap['from'], gap['to']])
            style_excel_sheet(ws, output_data['gaps'])
            wb.save(output_filename)
        elif args.format == 'json':
            with open(output_filename, 'w') as f:
                json.dump(output_data, f, indent=2)
        elif args.format == 'text':
            with open(output_filename, 'w') as f:
                for gap in output_data['gaps']:
                    f.write(args.pattern.format(**gap) + '\n')
    
    if args.output_dest in ['stdout', 'both']:
        if args.format == 'json':
            print(json.dumps(output_data, indent=2))
        elif args.format == 'text':
            for gap in output_data['gaps']:
                print(args.pattern.format(**gap))

def main():
    parser = argparse.ArgumentParser(
        description='Detect gaps in timestamp data from a CSV file.',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="""Example usage:
  # Excel output to file
  gap_detector.py --input data.csv --timestamp capture_time --format excel --output file
  
  # JSON output to stdout
  gap_detector.py --input data.csv --format json --output stdout
  
  # Text output with pattern to both file and stdout
  gap_detector.py --input data.csv --format text --pattern "Gap from {from} to {to}" --output both
  
  # With filename prefix
  gap_detector.py --input data.csv --prefix "analysis_" --format json --output file
""")
    
    parser.add_argument('--timestamp', default='_time', help='Name of the timestamp column')
    parser.add_argument('--delta', default='2h', type=parse_delta, 
                       help='Minimum gap duration (e.g., 2h, 30m, 1d)')
    parser.add_argument('--timezone', default='UTC', 
                       help='Timezone for output timestamps')
    parser.add_argument('--input', type=str, help='Path to input CSV file')
    parser.add_argument('--out', default='.', help='Output directory for files')
    parser.add_argument('--delimiter', default=';', help='CSV delimiter character')
    parser.add_argument('--format', default='excel', choices=['excel', 'json', 'text'],
                      help='Output format')
    parser.add_argument('--output', dest='output_dest', default='file',
                      choices=['file', 'stdout', 'both'], help='Output destination')
    parser.add_argument('--pattern', default='{from} -> {to}',
                      help='Pattern for text output (must contain {from} and/or {to})')
    parser.add_argument('--prefix', default='', help='Prefix for output filename')
    
    args = parser.parse_args()
    
    # Normalize output directory path
    args.out = os.path.normpath(args.out)
    
    # Read data from stdin or file
    if args.input:
        try:
            df = pd.read_csv(args.input, delimiter=args.delimiter)
        except Exception as e:
            logger.error(f"Error reading input file: {e}")
            sys.exit(1)
    else:
        if sys.stdin.isatty():
            parser.print_help()
            sys.exit(1)
        try:
            data = sys.stdin.read()
            df = pd.read_csv(StringIO(data), delimiter=args.delimiter)
        except Exception as e:
            logger.error(f"Error reading from stdin: {e}")
            sys.exit(1)
    
    # Check if timestamp column exists
    if args.timestamp not in df.columns:
        logger.error(f"Timestamp column '{args.timestamp}' not found in input data")
        sys.exit(1)
    
    # Convert to datetime
    try:
        df[args.timestamp] = pd.to_datetime(df[args.timestamp], utc=True)
    except Exception as e:
        logger.error(f"Error parsing timestamp column: {e}")
        sys.exit(1)
    
    # Detect gaps
    tz = pytz.timezone(args.timezone)
    gaps = detect_gaps(df, args.timestamp, args.delta, args.timezone)
    
    # Prepare output data
    from_time = df[args.timestamp].min().astimezone(tz).strftime('%Y-%m-%d_%H-%M-%S')
    to_time = df[args.timestamp].max().astimezone(tz).strftime('%Y-%m-%d_%H-%M-%S')
    iso_from = df[args.timestamp].min().astimezone(tz).isoformat()
    iso_to = df[args.timestamp].max().astimezone(tz).isoformat()
    
    output_data = {
        'period': {
            'from': iso_from,
            'to': iso_to
        },
        'gaps': gaps
    }
    
    if not gaps:
        logger.info("No gaps found in the data.")
        sys.exit(0)
    
    # Write output
    write_output(args, output_data, from_time, to_time)
    
    if args.output_dest in ['file', 'both'] and args.format == 'excel':
        logger.info(f"Gap report saved to {os.path.join(args.out, generate_output_filename(args, from_time, to_time))}")
    elif args.output_dest in ['file', 'both']:
        logger.info(f"Output saved to {os.path.join(args.out, generate_output_filename(args, from_time, to_time))}")

if __name__ == '__main__':
    main()