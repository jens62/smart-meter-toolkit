#!/usr/bin/env python3
"""Insert a "Lücken" (gaps) Start/End block into the Verbrauch sheet of a
consumption workbook produced by meter_reading2consumption.py, mirroring the
layout found in older reports (merged header, Start/End columns, zebra
striping) that used to be assembled by hand from gap_detector.py's output.
"""
import argparse
import json
import subprocess
import sys
import os

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter

GAP_DETECTOR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gap_detector.py")


def get_gaps(input_csv, timestamp, delimiter, delta, timezone):
    """Run gap_detector.py and return its list of {'from', 'to'} gaps (ISO, local tz)."""
    result = subprocess.run(
        [
            sys.executable, GAP_DETECTOR,
            "--input", input_csv,
            "--timestamp", timestamp,
            "--delimiter", delimiter,
            "--delta", delta,
            "--timezone", timezone,
            "--format", "json",
            "--output", "stdout",
        ],
        capture_output=True, text=True, check=True,
    )
    # gap_detector.py also prints a debug df.head() dump before the JSON
    json_start = result.stdout.index('{')
    data = json.loads(result.stdout[json_start:])
    return data["gaps"]


def add_gaps_to_sheet(xlsx_path, gaps, sheet_name, start_col):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    start_idx = column_index_from_string(start_col)
    end_idx = start_idx + 1
    duration_idx = start_idx + 2
    start_letter = start_col
    end_letter = get_column_letter(end_idx)
    duration_letter = get_column_letter(duration_idx)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    light_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
    dark_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    center = Alignment(horizontal="center")
    date_fmt = "dd.mm.yyyy hh:mm:ss"
    duration_fmt = "d hh:mm:ss"

    # Merged "Lücken" title over the three columns
    ws.merge_cells(f"{start_letter}1:{duration_letter}1")
    ws[f"{start_letter}1"] = "Lücken"
    for addr in (f"{start_letter}1", f"{end_letter}1", f"{duration_letter}1"):
        ws[addr].font = header_font
        ws[addr].fill = header_fill
        ws[addr].alignment = center

    # "Start" / "Ende" / "Dauer" sub-header
    ws.cell(row=2, column=start_idx, value="Start")
    ws.cell(row=2, column=end_idx, value="Ende")
    ws.cell(row=2, column=duration_idx, value="Dauer")
    for col in (start_idx, end_idx, duration_idx):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # Gap rows, zebra-striped like gap_detector.py's own output
    for i, gap in enumerate(gaps, start=2):
        row = i + 1
        fill = light_fill if i % 2 == 0 else dark_fill
        from_dt = pd.to_datetime(gap["from"]).tz_localize(None)
        to_dt = pd.to_datetime(gap["to"]).tz_localize(None)
        ws.cell(row=row, column=start_idx, value=from_dt)
        ws.cell(row=row, column=end_idx, value=to_dt)
        ws.cell(row=row, column=duration_idx, value=f"={end_letter}{row}-{start_letter}{row}")
        for col in (start_idx, end_idx):
            cell = ws.cell(row=row, column=col)
            cell.number_format = date_fmt
            cell.border = thin_border
            cell.alignment = center
            cell.fill = fill
        duration_cell = ws.cell(row=row, column=duration_idx)
        duration_cell.number_format = duration_fmt
        duration_cell.border = thin_border
        duration_cell.alignment = center
        duration_cell.fill = fill

    ws.column_dimensions[start_letter].width = 19
    ws.column_dimensions[end_letter].width = 19
    ws.column_dimensions[duration_letter].width = 14

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(
        description="Add a Lücken (gaps) Start/End block to the Verbrauch sheet of a workbook.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="""Example usage:
  add_gaps_to_verbrauch.py --xlsx report.xlsx --input normalized.csv --delta 20m
""",
    )
    parser.add_argument("--xlsx", required=True, help="Workbook to modify in place")
    parser.add_argument("--input", required=True, help="CSV with the readings used to detect gaps")
    parser.add_argument("--timestamp", default="_time", help="Name of the timestamp column")
    parser.add_argument("--delimiter", default=";", help="CSV delimiter character")
    parser.add_argument("--delta", default="20m", help="Minimum gap duration (e.g., 20m, 2h)")
    parser.add_argument("--timezone", default="Europe/Berlin", help="Timezone for displayed timestamps")
    parser.add_argument("--sheet", default="Verbrauch", help="Sheet to add the gaps block to")
    parser.add_argument("--start-col", default="F", help="Left column letter for the gaps block")
    args = parser.parse_args()

    gaps = get_gaps(args.input, args.timestamp, args.delimiter, args.delta, args.timezone)
    if not gaps:
        print("No gaps found - workbook left unchanged.")
        return

    add_gaps_to_sheet(args.xlsx, gaps, args.sheet, args.start_col)
    print(f"Added {len(gaps)} gap(s) to '{args.sheet}' sheet in {args.xlsx}")


if __name__ == "__main__":
    main()
