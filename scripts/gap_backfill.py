#!/usr/bin/env python3
"""Nightly gap-filling: detect gaps in one or more workbooks' recent months
and re-request each one from the gateway via read_SMGW.py.

The gateway itself caches roughly 15 months of readings (measured
2026-07-14 for this household's specific gateway - see TODO.md item 8 and
local-assets/smgw_retention_probe/ for how; re-measure if this matters
again much later or after a firmware update), so a gap caused by a missed
poll (e.g. the cron host being down) is usually still recoverable after
the fact - this is the "gap in already-exported data" technique from the
README's "Filling gaps in already-exported data" section, run
automatically instead of by hand.

Only the last `--months` month-sheets of each workbook are scanned
(default 15, matching the measured gateway retention above - see TODO.md
item 8), matching this script's purpose: recovering *recent,
still-recoverable* gaps while they're still in the gateway's own
retention window. Older gaps remain visible forever in the `--add-gaps`
"Luecken" block on the Verbrauch sheet - this script doesn't need to, and
deliberately doesn't, cover those. One consequence of the month-window
cut: a gap whose start falls in the month just before the window (i.e.
the boundary between the excluded and included range) is invisible here,
since gap detection only compares timestamps it was actually given.

Multi-meter handling: with --workbook-dir (the normal, cron-scheduled
mode), every *_from_*.xlsx workbook found there (latest-modified file per
distinct meter-prefix, in case older superseded copies are lying around)
is checked against the meters the gateway's own meter-select form
*currently* reports (read_SMGW.py --list-meters) - a workbook whose meter
isn't in that live list is skipped, since the gateway definitionally has
nothing to recover for a meter it no longer sees (e.g. after this
household's ITR03 -> EMH00 swap - though note the gateway kept *both* in
its list when this was checked 2026-07-14, so "no longer shown" isn't
guaranteed to happen even after a swap; don't assume it will).
--workbook and --meter together bypass discovery entirely, for manual
single-workbook use (e.g. testing) - both must be given, since guessing
one from the other isn't attempted.

A gap is retried once per calendar day (state persisted in --state-file,
so re-running this script multiple times in the same day is a no-op for
gaps already attempted today - safe for manual testing). After
--max-retries attempts across separate days with the gap still open, it's
marked "given up" and skipped on every later run, until the day it
actually disappears from detection (at which point it's dropped from the
state file as resolved). --max-requests-per-run is a shared budget across
every workbook processed in one run, not per-workbook.

This script never touches any workbook directly - it only writes new raw
export files into --out-path's data/ subdirectory, the same place
smgw2influx.sh's underlying reader writes them, and the same place
--append-to and daily-tar.sh already expect to find them. Picking up the
recovered data into a workbook and its Luecken block still happens via
the normal --append-to --add-gaps run, whenever that next runs.

If --influx-url/--influx-org/--influx-bucket/--influx-measurement (and
--influx-token or $INFLUXDB_TOKEN) are all given, a successfully-recovered
gap's readings are also written to InfluxDB - the regular smgw2influx.sh
polling job only ever writes "--past N minutes from now", so without this
a gap recovered here would never appear in InfluxDB at all, only in the
Excel pipeline. Uses the same line-protocol shape (line 77 of the legacy
smgw2influx.sh on the deployment host: "<measurement>,item=<measurement>
value=<v> <epoch_ns>") so recovered points land in the same series as
everything else, not a separate one. All discovered/processed meters
share the same --influx-measurement - if that's ever wrong for a
multi-meter household, this would need a per-meter measurement mapping,
which doesn't exist yet.
"""

import argparse
import glob
import json
import logging
import os
import subprocess
import sys
from datetime import date, datetime, timedelta, timezone

import pandas as pd
import pytz
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from meter_reading2consumption import MONTH_TAG_RE, find_gaps, format_measurement, read_month_rows  # noqa: E402

from openpyxl import load_workbook


def setup_logging(log_level):
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        stream=sys.stderr,
    )
    return logging.getLogger(__name__)


def gap_key(meter, gap_start, gap_end):
    return f"{meter}::{gap_start.isoformat()}__{gap_end.isoformat()}"


def collect_recent_times(xlsx_path, months, logger):
    """Sorted-descending datetimes from the workbook's last `months`
    month-sheets, plus the earliest month-sheet name included (so callers
    can tell whether a since-vanished gap aged out of the window or was
    actually resolved)."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        month_sheets = sorted(s for s in wb.sheetnames if MONTH_TAG_RE.match(s))
        if not month_sheets:
            raise ValueError(f"No monthly sheets found in {xlsx_path}")
        window = month_sheets[-months:]
        logger.info(f"Scanning {len(window)} month sheet(s) in {os.path.basename(xlsx_path)}: {', '.join(window)}")
        times = []
        for name in window:
            for dt, _time_str, _value, _meas in read_month_rows(wb[name]):
                times.append(dt)
        times.sort(reverse=True)
        window_start = datetime.strptime(f"{window[0]}_01", "%Y_%m_%d")
        return times, window_start
    finally:
        wb.close()


def workbook_own_meter(xlsx_path, logger):
    """The Zaehleridentifikation value stored in this workbook's own data
    (first data row of its earliest month-sheet) - the workbook's actual
    meter, read from its content rather than assumed from its filename.
    None if the workbook has no data rows at all."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        month_sheets = sorted(s for s in wb.sheetnames if MONTH_TAG_RE.match(s))
        if not month_sheets:
            return None
        rows = read_month_rows(wb[month_sheets[0]])
        return rows[0][3] if rows else None
    finally:
        wb.close()


def find_candidate_workbooks(workbook_dir, logger):
    """Latest-modified *_from_*.xlsx per distinct meter *actually found in
    each file's own data* - not per filename prefix. Filename prefixes
    aren't a reliable grouping key on their own: an old, abandoned
    workbook from an earlier pipeline version can use a different naming
    convention for the same meter (found directly on the deployment host -
    a stale "01005e31803c.1emh0011802881.sm_from_..." workbook for the same
    meter as the live "1_EMH00_1180_2881_from_..." one, last touched over a
    year before this was written) - grouping by filename would process
    that meter twice, once per naming convention, instead of once.

    Even grouping by the workbook's own stored id isn't quite enough on its
    own: that same old workbook stores the *raw* dotted logical name
    (format_measurement() hadn't been introduced yet when it was written),
    while the live workbook stores the already-formatted form - same
    meter, two different strings. format_measurement() is idempotent (a
    value that's already formatted passes through unchanged), so running
    every stored id through it first normalizes both to the same canonical
    key regardless of which convention a given workbook happens to use."""
    paths = glob.glob(os.path.join(workbook_dir, "*_from_*.xlsx"))
    by_meter = {}
    for p in paths:
        raw_id = workbook_own_meter(p, logger)
        if raw_id is None:
            logger.warning(f"{os.path.basename(p)}: no data rows, skipping")
            continue
        meter_id = format_measurement(raw_id)
        if meter_id not in by_meter or os.path.getmtime(p) > os.path.getmtime(by_meter[meter_id]):
            by_meter[meter_id] = p
    return sorted(by_meter.values())


def discover_live_meters(args, logger):
    """Raw logical names of every meter currently in the gateway's own
    meter-select form, via read_SMGW.py --list-meters. Empty list (not
    None) on a clean "gateway reports no meters" response; None on an
    actual failure to reach/parse the gateway."""
    cmd = [
        args.python, args.read_smgw_script,
        "--host", args.host,
        "--user", args.user,
        "--password", args.password,
        "--list-meters",
    ]
    logged_cmd = [c if c != args.password else "***" for c in cmd]
    logger.info(f"Discovering currently-connected meters: {' '.join(logged_cmd)}")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=args.request_timeout)
    except subprocess.TimeoutExpired:
        logger.error(f"--list-meters timed out after {args.request_timeout}s")
        return None
    if result.returncode != 0:
        logger.error(f"--list-meters failed: {result.stderr.strip()[-500:]}")
        return None
    return [line.strip() for line in result.stdout.splitlines() if line.strip()]


def load_state(path):
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        return json.load(f)


def save_state(path, state):
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w") as f:
        json.dump(state, f, indent=2, sort_keys=True)
    os.replace(tmp_path, path)


def pad_local_time(naive_local_dt, minutes, timezone):
    """naive_local_dt +/- minutes of real elapsed time, DST-safe.

    Plain `timedelta` arithmetic on a naive wall-clock time can land on a
    local time that doesn't exist (e.g. padding across the spring-forward
    night can produce "02:15", even though Europe/Berlin's clocks jump
    straight from 02:00 to 03:00 that night) or is ambiguous (the
    fall-back night's repeated hour). Localizing, adding in absolute time,
    then normalizing back gives the correct real-time-shifted wall clock
    instead - the same pytz idiom meter_reading2consumption.py's
    find_gaps() relies on for the same reason.
    """
    tz = pytz.timezone(timezone)
    aware = tz.localize(naive_local_dt)
    return tz.normalize(aware + timedelta(minutes=minutes)).replace(tzinfo=None)


def query_gateway(args, gap_start, gap_end, meter, logger):
    """Returns (ok, csv_path). csv_path is None if the request failed or
    the expected output file wasn't found (e.g. a chunking edge case)."""
    from_str = pad_local_time(gap_start, -args.pad_minutes, args.timezone).strftime("%Y-%m-%d %H:%M:%S")
    to_str = pad_local_time(gap_end, args.pad_minutes, args.timezone).strftime("%Y-%m-%d %H:%M:%S")

    cmd = [
        args.python, args.read_smgw_script,
        "--host", args.host,
        "--user", args.user,
        "--password", args.password,
        "--from", from_str,
        "--to", to_str,
        "--out-path", args.out_path,
    ]
    if meter:
        cmd += ["--meter", meter]

    logged_cmd = [c if c != args.password else "***" for c in cmd]
    logger.info(f"Querying gateway for {from_str} .. {to_str}: {' '.join(logged_cmd)}")

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=args.request_timeout)
    except subprocess.TimeoutExpired:
        logger.warning(f"Gateway request timed out after {args.request_timeout}s")
        return False, None

    if result.returncode != 0:
        logger.warning(
            f"read_SMGW.py exited {result.returncode} for {from_str} .. {to_str}: "
            f"{result.stderr.strip()[-500:]}"
        )
        return False, None

    # read_SMGW.py names a live query's combined output
    # "export_from_none-files_export_<from>---<to>.csv" (its generate_outputs()
    # prefixes with "export_from_<input_format>-files_" whenever source_files
    # is non-empty, which it always is for a live gateway query) - not the
    # plain "export_<from>---<to>.csv" the per-chunk .cms/.xml files use.
    # Glob on the escaped from/to substring rather than hardcoding that
    # prefix, so this doesn't silently break if that naming changes.
    escaped_from = from_str.replace(":", "_").replace(" ", "__")
    escaped_to = to_str.replace(":", "_").replace(" ", "__")
    pattern = os.path.join(args.out_path, "data", f"*{escaped_from}---{escaped_to}*.csv")
    matches = glob.glob(pattern)
    if not matches:
        logger.warning(f"read_SMGW.py succeeded but no output CSV found matching {pattern}")
        return True, None
    return True, max(matches, key=os.path.getmtime)


def influx_enabled(args):
    return bool(args.influx_url and args.influx_org and args.influx_bucket
                and args.influx_measurement and args.influx_token)


def write_to_influx(csv_path, args, logger):
    """Push a recovered gap's readings to InfluxDB, same line-protocol shape
    as the legacy smgw2influx.sh, so they land in the same series instead of
    a separate one. read_SMGW.py's CSV columns are:
    logical_name;capture_time;long64_value;scaler;unit;status;signature -
    note this differs from readSMGW_multipleContracts.sh's own CSV (used by
    smgw2influx.sh), which has value in column 2, not long64_value in
    column 3 - the column position matters here, not just the name."""
    lines = []
    with open(csv_path) as f:
        header = f.readline()
        if not header.startswith("logical_name;"):
            logger.warning(f"Unexpected CSV header in {csv_path}, skipping InfluxDB write: {header.strip()}")
            return 0
        for line in f:
            line = line.strip()
            if not line:
                continue
            _logical_name, capture_time, long64_value, _scaler, _unit, _status, _signature = line.split(";")
            epoch = int(datetime.strptime(capture_time, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc).timestamp())
            value = float(long64_value) / args.divisor
            lines.append(f"{args.influx_measurement},item={args.influx_measurement} value={value:.4f} {epoch}000000000")

    if not lines:
        logger.info(f"No rows in {csv_path} to write to InfluxDB")
        return 0

    resp = requests.post(
        f"{args.influx_url}/api/v2/write?org={args.influx_org}&bucket={args.influx_bucket}",
        headers={"Authorization": f"Token {args.influx_token}"},
        data="\n".join(lines),
        timeout=30,
    )
    if resp.status_code // 100 != 2:
        logger.warning(f"InfluxDB write failed ({resp.status_code}): {resp.text[:300]}")
        return 0
    logger.info(f"Wrote {len(lines)} row(s) to InfluxDB from {os.path.basename(csv_path)}")
    return len(lines)


def resolve_workbook_meter_pairs(args, logger):
    """[(workbook_path, meter), ...] to actually process this run."""
    if args.workbook and args.meter:
        return [(args.workbook, args.meter)]
    if args.workbook or args.meter:
        logger.error("--workbook and --meter must be given together (or neither, to use --workbook-dir discovery)")
        sys.exit(1)
    if not args.workbook_dir:
        logger.error("Either --workbook and --meter together, or --workbook-dir, must be given")
        sys.exit(1)

    live_meters = discover_live_meters(args, logger)
    if live_meters is None:
        sys.exit(1)
    logger.info(f"Gateway currently shows {len(live_meters)} meter(s): {', '.join(live_meters) or '(none)'}")

    candidates = find_candidate_workbooks(args.workbook_dir, logger)
    logger.info(f"Found {len(candidates)} candidate workbook(s) in {args.workbook_dir}")

    pairs = []
    for wb_path in candidates:
        raw_stored_id = workbook_own_meter(wb_path, logger)
        if raw_stored_id is None:
            logger.warning(f"{os.path.basename(wb_path)}: no data rows, skipping")
            continue
        stored_id = format_measurement(raw_stored_id)  # normalize old-style raw ids too, see find_candidate_workbooks()
        match = next((m for m in live_meters if format_measurement(m) == stored_id), None)
        if match is None:
            logger.info(
                f"{os.path.basename(wb_path)} (meter '{stored_id}') is not currently connected "
                f"to the gateway - skipping, nothing recoverable"
            )
            continue
        pairs.append((wb_path, match))
    return pairs


def process_workbook(workbook_path, meter, args, state, today, budget_remaining, logger):
    """One workbook's worth of detect-gaps-and-recover. Mutates `state` in
    place. Returns (queried, failed, influx_written, deferred_count,
    new_budget_remaining)."""
    delta = pd.Timedelta(args.delta)

    try:
        times, window_start = collect_recent_times(workbook_path, args.months, logger)
    except (FileNotFoundError, ValueError) as e:
        logger.error(str(e))
        return 0, 0, 0, 0, budget_remaining

    detected = find_gaps(times, threshold=delta, timezone=args.timezone)
    detected.sort(key=lambda g: g[0])  # oldest gap first - find_gaps() returns newest-first
    detected_keys = {gap_key(meter, start, end): (start, end) for start, end, _duration in detected}
    logger.info(f"{meter}: detected {len(detected)} gap(s) over {delta} in the last {args.months} month(s)")

    for key in list(state.keys()):
        if key.startswith(f"{meter}::") and key not in detected_keys:
            entry = state.pop(key)
            gap_end = datetime.fromisoformat(entry["end"])
            if gap_end < window_start:
                logger.info(f"{meter}: gap {entry['start']} .. {entry['end']} aged out of the {args.months}-month window (untracked)")
            else:
                logger.info(f"{meter}: gap {entry['start']} .. {entry['end']} resolved (no longer detected)")

    to_query = []
    for key, (start, end) in detected_keys.items():
        entry = state.setdefault(key, {
            "meter": meter, "start": start.isoformat(), "end": end.isoformat(),
            "attempts": 0, "last_attempt_date": None, "status": "open",
        })
        if entry["status"] == "given_up":
            logger.debug(f"{meter}: gap {start} .. {end}: given up previously, skipping")
            continue
        if entry["attempts"] >= args.max_retries:
            entry["status"] = "given_up"
            logger.warning(f"{meter}: gap {start} .. {end}: still open after {entry['attempts']} attempts, giving up")
            continue
        if entry["last_attempt_date"] == today:
            logger.debug(f"{meter}: gap {start} .. {end}: already attempted today, skipping")
            continue
        to_query.append((key, start, end, entry))

    deferred = to_query[budget_remaining:]
    to_query = to_query[:budget_remaining]
    if deferred:
        logger.warning(
            f"{meter}: deferring {len(deferred)} eligible gap(s) to a later run (shared --max-requests-per-run budget exhausted): "
            + ", ".join(f"{s} .. {e}" for _k, s, e, _entry in deferred)
        )

    queried, failed, influx_written = 0, 0, 0
    for key, start, end, entry in to_query:
        if args.dry_run:
            logger.info(f"{meter}: [dry-run] would query gateway for gap {start} .. {end} (attempt {entry['attempts'] + 1}/{args.max_retries})")
            continue
        ok, csv_path = query_gateway(args, start, end, meter, logger)
        entry["attempts"] += 1
        entry["last_attempt_date"] = today
        entry["last_result"] = "ok" if ok else "failed"
        queried += 1
        if not ok:
            failed += 1
        elif csv_path and influx_enabled(args):
            influx_written += write_to_influx(csv_path, args, logger)

    return queried, failed, influx_written, len(deferred), budget_remaining - len(to_query)


def main():
    parser = argparse.ArgumentParser(
        description="Detect recent gaps in one or more workbooks and re-request them from the gateway.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # Normal, cron-scheduled mode: discover every currently-connected meter
  # and match it against workbooks found in --workbook-dir.
  %(prog)s \\
      --workbook-dir /path/to/workbook-directory \\
      --state-file /path/to/gap_backfill_state.json \\
      --out-path /path/to/data-directory \\
      --user <user> --password <password> \\
      --dry-run

  # Manual single-workbook override, bypassing discovery entirely:
  %(prog)s \\
      --workbook /path/to/1_EMH00_..._to_....xlsx --meter <meter> \\
      --state-file /path/to/gap_backfill_state.json \\
      --out-path /path/to/data-directory \\
      --user <user> --password <password> \\
      --dry-run
""",
    )
    parser.add_argument("--workbook-dir", help="Directory to discover workbooks in (normal mode)")
    parser.add_argument("--workbook", help="Single workbook to scan - requires --meter too, bypasses discovery")
    parser.add_argument("--meter", help="Meter id for --workbook - requires --workbook too, bypasses discovery")
    parser.add_argument("--months", type=int, default=15,
                        help="How many trailing month-sheets to scan (default: 15, matching this "
                             "gateway's measured retention window - see TODO.md item 8)")
    parser.add_argument("--delta", default="20m", help="Minimum gap duration to act on (e.g. 20m, 1h)")
    parser.add_argument("--timezone", default="Europe/Berlin", help="Timezone for gap detection")

    parser.add_argument("--state-file", required=True, help="JSON file tracking per-gap retry state")
    parser.add_argument("--max-retries", type=int, default=3,
                        help="Give up on a gap after this many days of still being open")
    parser.add_argument("--max-requests-per-run", type=int, default=10,
                        help="Cap gateway requests in one run, shared across every workbook processed "
                             "(be gentle on the gateway); excess eligible gaps are deferred to the "
                             "next run, oldest first")
    parser.add_argument("--pad-minutes", type=int, default=30,
                        help="Widen each gateway request by this many minutes on either side "
                             "of the detected gap, to be safe against boundary readings")
    parser.add_argument("--request-timeout", type=int, default=180,
                        help="Per-request subprocess timeout in seconds")

    parser.add_argument("--host", default="192.168.1.200", help="Gateway IP")
    parser.add_argument("--user", required=True, help="Gateway user")
    parser.add_argument("--password", default=os.environ.get("SMGW_PASSWORD"),
                        help="Gateway password (default: $SMGW_PASSWORD)")
    parser.add_argument("--out-path", required=True,
                        help="Base directory read_SMGW.py writes into (creates/uses its own data/ subdir)")
    parser.add_argument("--read-smgw-script",
                        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "read_SMGW.py"),
                        help="Path to read_SMGW.py")
    parser.add_argument("--python", default=sys.executable, help="Python interpreter to run read_SMGW.py with")

    parser.add_argument("--divisor", type=float, default=10000.0,
                        help="Raw counter value -> InfluxDB unit, e.g. Wh -> kWh (default: 10000)")
    parser.add_argument("--influx-url", help="InfluxDB base URL, e.g. http://host:8086 - omit to skip InfluxDB entirely")
    parser.add_argument("--influx-org", help="InfluxDB org")
    parser.add_argument("--influx-bucket", help="InfluxDB bucket")
    parser.add_argument("--influx-measurement", help="InfluxDB measurement/item name, e.g. SMGW_EPPC0211923304")
    parser.add_argument("--influx-token", default=os.environ.get("INFLUXDB_TOKEN"),
                        help="InfluxDB API token (default: $INFLUXDB_TOKEN)")

    parser.add_argument("--dry-run", action="store_true",
                        help="Detect and report gaps/state changes without querying the gateway "
                             "or writing the state file")
    parser.add_argument("--log-level", default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
    args = parser.parse_args()

    logger = setup_logging(args.log_level)

    if not args.password:
        logger.error("No gateway password given (--password or $SMGW_PASSWORD)")
        sys.exit(1)

    if influx_enabled(args):
        logger.info(f"InfluxDB writes enabled: {args.influx_url} org={args.influx_org} bucket={args.influx_bucket} measurement={args.influx_measurement}")
    else:
        logger.info("InfluxDB writes disabled (not all of --influx-url/--influx-org/--influx-bucket/--influx-measurement/--influx-token given)")

    pairs = resolve_workbook_meter_pairs(args, logger)
    if not pairs:
        logger.info("No workbook to process this run (none matched a currently-connected meter) - nothing to do")
        sys.exit(0)

    state = load_state(args.state_file)
    today = date.today().isoformat()
    budget = args.max_requests_per_run

    total_detected_workbooks = len(pairs)
    total_queried = total_failed = total_influx_written = total_deferred = 0
    for workbook_path, meter in pairs:
        queried, failed, influx_written, deferred, budget = process_workbook(
            workbook_path, meter, args, state, today, budget, logger
        )
        total_queried += queried
        total_failed += failed
        total_influx_written += influx_written
        total_deferred += deferred

    if not args.dry_run:
        save_state(args.state_file, state)
    else:
        logger.info("[dry-run] state file not written")

    logger.info(
        f"Done: {total_detected_workbooks} workbook(s) processed, {total_queried} queried ({total_failed} failed), "
        f"{total_influx_written} row(s) written to InfluxDB, "
        f"{total_deferred} deferred, "
        f"{sum(1 for e in state.values() if e['status'] == 'given_up')} given up total"
    )


if __name__ == "__main__":
    main()
