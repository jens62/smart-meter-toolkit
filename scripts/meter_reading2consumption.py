#!/usr/bin/env python3

import os
import sys
import re
import glob
import io
import subprocess
import logging
import pandas as pd
import argparse
from datetime import datetime
import pytz
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

NUMBER_FORMAT = '#,##0.0000;[Red]-#,##0.0000'

NORMALIZE_AWK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "normalize_meter_csv.awk")

def setup_logging(log_level=logging.INFO):
    """Configure logging with custom formatter and specified log level."""
    logger = logging.getLogger()
    logger.setLevel(log_level)
    
    ch = logging.StreamHandler()
    ch.setLevel(log_level)
    
    formatter = logging.Formatter(
        '%(asctime)-15s %(name)-8s %(lineno)d %(levelname)s: %(message)s'
    )
    ch.setFormatter(formatter)
    
    if not logger.handlers:
        logger.addHandler(ch)
    
    return logger

# Extract part between dots and format it, if length is 14 chars, according to DIN 43863-5
# see https://netze.estw.de/erlangenGips/Erlangen/__attic__20210120_155237__estw1.de/Kopfnavigation/Netze/Messwesen/Messwesen/Herstelleruebergreifende-Identifikationsnummer-fuer-Messeinrichtungen.pdf
# Zeichen 1: Spartenkennung
# Zeichen 2-4: Herstellerkennzeichnung
# Zeichen 5-6: Fabrikationsblock
# Zeichen 7-14: Fabrikationsnummer
#
# Kennung;Sparte;Erläuterung
# 0;–;Die 0 ist wegen unterschiedlicher Darstellung und Verwendung in den Geräteverwaltungssystemen nicht zu verwenden.
# 1;Elektrizität;     
# 2;–;  
# 3;–;   
# 4;Heizkosten;  
# 5;Kälte;   
# 6;Wärme;   
# 7;Gas; 
# 8;Wasser, kalt;Temperatur Medium < 30 °C  
# 9;Wasser, heiß;Temperatur Medium 30 °C … 90 °C und 90 °C  
# A;–    ;
# B;–    ;
# C;–    ;
# D;–    ;
# E;    Kommunikation;Kommunikationsgeräte wie z. B. Datensammler stellen eine eigene Sparte dar und sind daher mit einer eigenen Kennung zu versehen.  
# F;    Bisher nicht spezifizierte Sparten;Um eine Konvertierung der Sparten nach OBIS zu anderen Kodierungen zu ermöglichen, wird der Buchstabe F als „Jokerzeichen“ für hier nicht weiter aufgeführte Sparten verwendet.
def format_measurement(x):
    parts = x.split(".")
    if len(parts) > 1 and len(parts[1]) == 14:
        return re.sub(r'^(.{1})(.{3})(.{2})(.{4})(.{4})$', 
                     r'\1 \2\3 \4 \5', 
                     parts[1]).upper()
    return x.rsplit(".", 1)[0]

def convert_to_timezone(df, timezone, logger=None):
    """Convert dataframe timestamps to specified timezone."""
    if not isinstance(timezone, str):
        raise ValueError(f"Timezone must be string (e.g. 'Europe/Berlin'), got {type(timezone)}")
    
    if logger and logger.level <= logging.DEBUG:
        logger.debug(f"Converting timestamps to timezone: {timezone}")
    
    # Ensure timestamps are timezone-aware (UTC)
    if df["_time"].dt.tz is None:
        df["_time"] = df["_time"].dt.tz_localize('UTC')
    
    # Convert to target timezone
    tz = pytz.timezone(timezone)
    df["_time"] = df["_time"].dt.tz_convert(tz)
    return df

def process_dataframe(df, keep_every, timezone, divisor, logger=None):
    """Apply common processing steps to a dataframe."""
    try:
        # 0. Normalize timestamps by removing fractions of seconds while preserving timezone
        if pd.api.types.is_datetime64_any_dtype(df['_time']):
            # For already parsed timestamps
            if df['_time'].dt.tz is None:
                # If no timezone, assume UTC
                df['_time'] = pd.to_datetime(df['_time']).dt.tz_localize('UTC').dt.floor('s')
            else:
                # If timezone exists, keep it
                df['_time'] = df['_time'].dt.floor('s')
        else:
            # For string timestamps, parse while preserving timezone info
            df['_time'] = pd.to_datetime(
                df['_time'].str.replace(r'(\.\d+)(?=[Z+-])', '', regex=True),  # Remove fractions before timezone
                utc=True  # Parse timezone-aware
            ).dt.floor('s')

        # 0.5 Apply downsampling if requested
        if keep_every != 1:
            if logger:
                logger.info(f"Downsampling data (keeping every {keep_every if keep_every > 0 else 'first/last per month'})")
            df = downsample_data(df, keep_every, timezone, logger)

        # 1. Format measurement column
        df['_measurement'] = df['_measurement'].apply(format_measurement)

        # 2. Sort by _time in descending order (youngest first)
        df = df.sort_values('_time', ascending=False)

        # 3. Remove duplicate rows (keeping the first occurrence - which will be the youngest due to sorting)
        df = df.drop_duplicates()

        # 4. Apply divisor if needed
        if divisor != 1:
            df['_value'] = df['_value'] / divisor

        # Ensure required columns exist and in correct order
        if not all(col in df.columns for col in ['_time', '_value', '_measurement']):
            raise ValueError("Missing required columns in input data")
    
        return df[['_time', '_value', '_measurement']]
    
    except Exception as e:
        if logger:
            logger.error(f"Error processing dataframe: {str(e)}")
        raise    

def load_from_stdin(time_col, value_col, measurement_col, delimiter, keep_every, timezone, divisor, logger=None):
    """Load and process data from stdin."""
    if logger:
        logger.info("Reading data from stdin...")
    
    try:
        df = pd.read_csv(sys.stdin, parse_dates=[time_col], delimiter=delimiter)
        df = df.rename(columns={
            time_col: '_time',
            value_col: '_value',
            measurement_col: '_measurement'
        })
        
        return process_dataframe(df, keep_every, timezone, divisor, logger)
    
    except Exception as e:
        if logger:
            logger.error(f"Error reading from stdin: {str(e)}")
        raise

def downsample_data(df, keep_every_n, timezone, logger=None):
    """
    Downsample data while preserving:
    - Original timestamps
    - Always keeps newest/oldest per timezone-based year-month
    - Maintains measurement boundaries
    """
    if keep_every_n <= 1:
        return df

    try:
        # Create timezone-aware copy just for grouping
        df_group = df.copy()
        df_group = convert_to_timezone(df_group, timezone, logger)
        df_group['_year_month'] = df_group['_time'].dt.strftime('%Y-%m')
        
        # Track indices to keep from original df
        keep_indices = set()
        
        for (year_month, measurement), group_indices in df_group.groupby(
            ['_year_month', '_measurement']).groups.items():
            
            group_size = len(group_indices)
            
            # Always keep newest and oldest
            keep_indices.add(group_indices[0])  # Newest (after sort)
            keep_indices.add(group_indices[-1]) # Oldest
            
            # Add sampled points between them
            if group_size > 2 and keep_every_n > 1:
                keep_indices.update(group_indices[1:-1:keep_every_n])
        
        # Return filtered original dataframe (with original timestamps)
        return df.iloc[sorted(keep_indices)].copy()
    
    except Exception as e:
        if logger:
            logger.error(f"Downsampling failed: {str(e)}")
        raise

def load_and_process_data(input_path, file_pattern, timezone, column_mapping, keep_every, divisor, logger=None, delimiter=','):
    """Load and process data from file(s)."""
    if logger:
        logger.info(f"Processing data with timezone: {timezone}")
    
    if os.path.isfile(input_path):
        df = pd.read_csv(input_path, parse_dates=[column_mapping['_time']], delimiter=delimiter)
    else:
        data_frames = []
        pattern = re.compile(file_pattern)
        
        for file in os.listdir(input_path):
            if pattern.search(file):
                try:
                    df = pd.read_csv(os.path.join(input_path, file), 
                                   parse_dates=[column_mapping['_time']], 
                                   delimiter=delimiter)
                    data_frames.append(df)
                except Exception as e:
                    if logger:
                        logger.error(f"Error processing file {file}: {str(e)}")
                    continue
        
        if not data_frames:
            raise ValueError(f"No files matching pattern '{file_pattern}' found")
        df = pd.concat(data_frames, ignore_index=True)
    
    # Rename columns
    df = df.rename(columns={v: k for k, v in column_mapping.items()})
    
    try:
        df = process_dataframe(df, keep_every, timezone, divisor, logger)
        df = convert_to_timezone(df, timezone, logger)
        df['_time'] = df['_time'].dt.tz_localize(None)  # Remove tz for processing
        return df
    except Exception as e:
        if logger:
            logger.error(f"Error processing data: {str(e)}")
        raise

def apply_german_number_format(ws, value_col_idx):
    """Apply German number formatting to value column"""
    german_number_format = '#,##0.0000;[Red]-#,##0.0000'
    for row in ws.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # cell.value = round(cell.value, 4)  # Ensure 4 decimal places
                cell.number_format = german_number_format

def apply_zebra_formatting(ws, header_style):
    """Apply zebra striping with borders to all cells"""
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            # Header row
            if cell.row == 1:
                for attr, value in header_style.items():
                    setattr(cell, attr, value)
            # Data rows
            else:
                fill = light_fill if cell.row % 2 == 0 else white_fill
                cell.fill = fill
                cell.border = header_style['border']

def set_dynamic_column_widths(ws, df, is_consumption=False):
    """Set optimized column widths including header text"""
    if is_consumption:
        # Verbrauch sheet columns
        header_lengths = {
            col: len(ws[f'{col}1'].value)
            for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G')
        }
        content_lengths = {
            'A': max(len(str(v.value)) for v in ws['A']),
            'B': max(len(f"{v.value:,.4f}") if isinstance(v.value, (int, float)) else len(str(v.value))
                   for v in ws['B'])
        }

        # Use whichever is longer - header or content
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 15)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)

        # C-G hold formulas; their computed values aren't known until Excel opens
        # the file, so size them from the header text plus headroom for the
        # eventual date/time or number display.
        for col, min_width in (('C', 20), ('D', 15), ('E', 20), ('F', 15), ('G', 15)):
            ws.column_dimensions[col].width = min(max(header_lengths[col], min_width) + 2, 35)

    else:
        # Monthly sheet columns
        headers = ["Zeit der Messung", f"Zählerstand [kWh]", "Zähleridentifikation"]
        header_lengths = {
            'A': len(headers[0]),
            'B': len(headers[1]),
            'C': len(headers[2])
        }
        
        content_lengths = {
            'A': max(len(row[df.columns.get_loc('_time')].strftime("%d.%m.%Y %H:%M:%S")) for row in df.itertuples(index=False)),
            'B': max(len(f"{v:,.4f}") if isinstance(v, (int, float)) else len(str(v)) 
                   for v in df["_value"]),
            'C': max(len(str(v)) for v in df["_measurement"])
        }
        
        # Apply widths with header consideration
        ws.column_dimensions['A'].width = min(max(header_lengths['A'], content_lengths['A']) + 2, 22)
        ws.column_dimensions['B'].width = min(max(header_lengths['B'], content_lengths['B']) + 2, 20)
        ws.column_dimensions['C'].width = min(max(header_lengths['C'], content_lengths['C']) + 2, 35)

def build_header_style():
    """Style dict shared by the initial-generation and append code paths."""
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
    }


def build_month_sheet(wb, ws_consumption, year_month, month_df, unit, header_style, row_num=None):
    """Create one monthly sheet and write its Verbrauch summary row.

    Shared by generate_excel_output() (full regeneration) and
    append_to_workbook() (adding newer months to an existing workbook), so
    the sheet layout and the Verbrauch formulas stay identical either way.

    row_num, if given, is the exact Verbrauch row to write into (used by
    append_to_workbook() when inserting ahead of hand-added rows below the
    last month, e.g. a "Summe" total row). Defaults to appending at the end.
    """
    number_format = NUMBER_FORMAT
    datetime_format = 'dd.mm.yyyy hh:mm:ss'

    # Create worksheet with German headers
    ws = wb.create_sheet(title=year_month)
    ws.append(["Zeit der Messung", f"Zählerstand [{unit}]", "Zähleridentifikation"])

    # Get column indices once
    time_idx = month_df.columns.get_loc('_time')
    value_idx = month_df.columns.get_loc('_value')
    meas_idx = month_df.columns.get_loc('_measurement')

    # Add data rows using column indices
    for row in month_df.itertuples(index=False):
        ws.append([
            row[time_idx].strftime("%d.%m.%Y %H:%M:%S"),
            row[value_idx],
            row[meas_idx]
        ])

    # Apply German number formatting
    apply_german_number_format(ws, value_col_idx=2)

    # Apply zebra formatting with borders
    apply_zebra_formatting(ws, header_style)

    # Set dynamic column widths
    set_dynamic_column_widths(ws, month_df)

    # Add consumption row. Rather than a naive MAX-MIN per month (which
    # drops the reading interval between two months' sheets, and would
    # be wrong anyway if the meter's value isn't strictly monotonic,
    # e.g. with feed-in/production), this chains each month's boundary
    # value to the next: the value is looked up at the exact moment of
    # the month boundary, linearly interpolated between this month's
    # last reading and next month's first reading if they don't land
    # exactly on the boundary. process_dataframe() sorts newest-first,
    # so a month's own latest reading is its first data row (row 2)
    # and its own oldest reading is its last data row.
    if row_num is None:
        row_num = ws_consumption.max_row + 1
    next_row = row_num + 1

    # NOTE: openpyxl writes formula text straight into the workbook's
    # XML, which always uses the canonical English function names and
    # comma argument separators, regardless of the user's Excel
    # locale. Excel translates this to the display locale (e.g.
    # WENN/ISTLEER + semicolons in German Excel) automatically when
    # the file is opened - writing localized names/separators here
    # instead produces a file Excel flags as corrupted.
    letzte_zeit = (
        f"""=DATEVALUE(INDIRECT("'"&A{row_num}&"'!A2"))"""
        f"""+TIMEVALUE(INDIRECT("'"&A{row_num}&"'!A2"))"""
    )
    letzter_wert = f"""=INDIRECT("'"&A{row_num}&"'!B2")"""
    erste_zeit_folgemonat = (
        f"""=IF(ISBLANK(A{next_row}),"","""
        f"""DATEVALUE(INDEX(INDIRECT("'"&A{next_row}&"'!A:A"),COUNTA(INDIRECT("'"&A{next_row}&"'!A:A"))))"""
        f"""+TIMEVALUE(INDEX(INDIRECT("'"&A{next_row}&"'!A:A"),COUNTA(INDIRECT("'"&A{next_row}&"'!A:A")))))"""
    )
    erster_wert_folgemonat = (
        f"""=IF(ISBLANK(A{next_row}),"","""
        f"""INDEX(INDIRECT("'"&A{next_row}&"'!B:B"),COUNTA(INDIRECT("'"&A{next_row}&"'!A:A"))))"""
    )
    grenzwert = (
        f"""=IF(ISBLANK(A{next_row}),D{row_num},"""
        f"""D{row_num}+(F{row_num}-D{row_num})*"""
        f"""(DATE(VALUE(LEFT(A{next_row},4)),VALUE(RIGHT(A{next_row},2)),1)-C{row_num})"""
        f"""/(E{row_num}-C{row_num}))"""
    )
    if row_num == 2:
        # Oldest month has no predecessor: use its own earliest reading.
        verbrauch = (
            f"""=G{row_num}-INDEX(INDIRECT("'"&A{row_num}&"'!B:B"),"""
            f"""COUNTA(INDIRECT("'"&A{row_num}&"'!A:A")))"""
        )
    else:
        verbrauch = f"=G{row_num}-G{row_num - 1}"

    for col, value in enumerate(
        (year_month, verbrauch, letzte_zeit, letzter_wert,
         erste_zeit_folgemonat, erster_wert_folgemonat, grenzwert),
        start=1,
    ):
        ws_consumption.cell(row=row_num, column=col, value=value)

    # Apply number formats to the newly added row's formula cells
    # (openpyxl doesn't evaluate formulas, so this can't rely on the
    # isinstance-based apply_german_number_format() helper)
    for col, fmt in (
        (2, number_format), (3, datetime_format), (4, number_format),
        (5, datetime_format), (6, number_format), (7, number_format),
    ):
        ws_consumption.cell(row=row_num, column=col).number_format = fmt


PLAUSIBILITAETSTEST_LABEL = "Plausibilitätstest: jüngster Wert - ältester Wert"


def add_summe_and_plausibility_rows(ws_consumption, last_month_row, header_style):
    """Append the fixed trailing block below the last month row: a blank
    spacer row, a "Summe" total, and a "Plausibilitaetstest" cross-check.

    The blank row is what keeps the last month's own ISBLANK(A{next_row})
    "is there a next month" check working unchanged - it lands on the blank
    row, not on "Summe", so no ISBLANK/ISREF rewrite is needed there.

    The Plausibilitaetstest formula (newest overall reading minus oldest
    overall reading) is a pure telescoping-sum identity: it must always
    equal Summe exactly if the per-month boundary-chaining formulas
    (generate_excel_output()'s "Verbrauch" column) are correct, making it a
    built-in regression check on those formulas rather than a data check.
    """
    summe_row = last_month_row + 2
    plausi_row = summe_row + 1

    ws_consumption.cell(row=summe_row, column=1, value="Summe")
    ws_consumption.cell(row=summe_row, column=2, value=f"=SUM(B2:B{last_month_row})")
    for col in (1, 2):
        cell = ws_consumption.cell(row=summe_row, column=col)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
    ws_consumption.cell(row=summe_row, column=2).number_format = NUMBER_FORMAT

    ws_consumption.cell(row=plausi_row, column=1, value=PLAUSIBILITAETSTEST_LABEL)
    ws_consumption.cell(row=plausi_row, column=2, value=(
        f"""=INDIRECT("'"&A{last_month_row}&"'!B2")-"""
        f"""INDEX(INDIRECT("'"&A2&"'!B:B"),COUNTA(INDIRECT("'"&A2&"'!A:A")))"""
    ))
    ws_consumption.cell(row=plausi_row, column=2).number_format = NUMBER_FORMAT


def add_gaps_block(ws_consumption, gaps, start_col='I', logger=None):
    """Write a "Luecken" Start/Ende/Dauer block into the Verbrauch sheet.

    Folds generate_excel/add_gaps_to_verbrauch.py's layout directly into
    the main generation path (its detection logic is find_gaps(), already
    shared with append_to_workbook() rather than duplicated) instead of it
    being a separate post-processing script/step.
    """
    if not gaps:
        if logger:
            logger.info("No gaps found - Luecken block left empty")
        return

    # find_gaps() walks its input newest-first, so its output comes out
    # newest-gap-first too - re-sort oldest-first (ascending by start) for
    # display, matching add_gaps_to_verbrauch.py's original ordering.
    gaps = sorted(gaps, key=lambda g: g[0])

    start_idx = column_index_from_string(start_col)
    end_idx = start_idx + 1
    duration_idx = start_idx + 2
    start_letter = start_col
    end_letter = get_column_letter(end_idx)
    duration_letter = get_column_letter(duration_idx)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    light_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
    dark_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    center = Alignment(horizontal="center")
    date_fmt = "dd.mm.yyyy hh:mm:ss"
    duration_fmt = "d hh:mm:ss"

    ws_consumption.merge_cells(f"{start_letter}1:{duration_letter}1")
    ws_consumption[f"{start_letter}1"] = "Lücken (keine Daten vom SMGW)"
    for addr in (f"{start_letter}1", f"{end_letter}1", f"{duration_letter}1"):
        ws_consumption[addr].font = header_font
        ws_consumption[addr].fill = header_fill
        ws_consumption[addr].alignment = center

    ws_consumption.cell(row=2, column=start_idx, value="Start")
    ws_consumption.cell(row=2, column=end_idx, value="Ende")
    ws_consumption.cell(row=2, column=duration_idx, value="Dauer")
    for col in (start_idx, end_idx, duration_idx):
        cell = ws_consumption.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for i, (gap_start, gap_end, _duration) in enumerate(gaps, start=2):
        row = i + 1
        fill = light_fill if i % 2 == 0 else dark_fill
        ws_consumption.cell(row=row, column=start_idx, value=gap_start)
        ws_consumption.cell(row=row, column=end_idx, value=gap_end)
        ws_consumption.cell(row=row, column=duration_idx, value=f"={end_letter}{row}-{start_letter}{row}")
        for col in (start_idx, end_idx):
            cell = ws_consumption.cell(row=row, column=col)
            cell.number_format = date_fmt
            cell.border = thin_border
            cell.alignment = center
            cell.fill = fill
        duration_cell = ws_consumption.cell(row=row, column=duration_idx)
        duration_cell.number_format = duration_fmt
        duration_cell.border = thin_border
        duration_cell.alignment = center
        duration_cell.fill = fill

    ws_consumption.column_dimensions[start_letter].width = 19
    ws_consumption.column_dimensions[end_letter].width = 19
    ws_consumption.column_dimensions[duration_letter].width = 14

    if logger:
        logger.info(f"Added {len(gaps)} gap(s) to the Luecken block")


def generate_excel_output(df, output_dir, measurement_name, unit, timezone, logger=None, add_gaps=False):
    """Generate Excel output with professional German formatting."""
    try:
        if logger:
            logger.info("Generating Excel output...")

        time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
        output_filename = f"{measurement_name}_{time_range}.xlsx"

        # df['_time'] is already naive Europe/Berlin local time at this point
        df_excel = df.copy()

        wb = Workbook()
        wb.remove(wb.active)

        header_style = build_header_style()

        # ===== VERBRAUCH SHEET =====
        ws_consumption = wb.create_sheet(title="Verbrauch")
        ws_consumption.append([
            "Monat",
            f"Verbrauch [{unit}]",
            "Letzte Messung [Zeit]",
            f"Letzte Messung [{unit}]",
            "Erste Messung Folgemonat [Zeit]",
            f"Erste Messung Folgemonat [{unit}]",
            f"Grenzwert Monatsende [{unit}]",
        ])

        # Apply header style
        for cell in ws_consumption[1]:
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # ===== MONTHLY SHEETS =====
        df_excel["_year_month"] = df_excel["_time"].dt.strftime("%Y_%m")
        for year_month, month_df in df_excel.groupby("_year_month"):
            build_month_sheet(wb, ws_consumption, year_month, month_df, unit, header_style)

        # ===== FINALIZE VERBRAUCH SHEET =====
        # Runs before the Summe/Plausibilitaetstest/Luecken blocks below,
        # since apply_zebra_formatting() blanket-restripes every row/column
        # already in the sheet's used range - adding those blocks first
        # would just get their own custom colors immediately overwritten.
        apply_zebra_formatting(ws_consumption, header_style)
        set_dynamic_column_widths(ws_consumption, df_excel, is_consumption=True)

        # ===== SUMME / PLAUSIBILITAETSTEST =====
        last_month_row = ws_consumption.max_row
        add_summe_and_plausibility_rows(ws_consumption, last_month_row, header_style)

        # ===== LUECKEN (optional) =====
        if add_gaps:
            gaps = find_gaps(sorted(df_excel['_time'].tolist(), reverse=True), timezone=timezone)
            add_gaps_block(ws_consumption, gaps, logger=logger)

        # Save workbook
        wb.save(os.path.join(output_dir, output_filename))
        return output_filename

    except Exception as e:
        if logger:
            logger.error(f"Error generating Excel output: {str(e)}")
        raise

def generate_csv_output(df, measurement_name, unit, timezone, logger=None):
    """Generate CSV formatted output with period and consumption data."""
    try:
        if logger:
            logger.debug("Generating CSV output with period data")
        
        # df['_time'] is already naive Europe/Berlin local time at this point
        df_csv = df.copy()

        # Add year and month columns
        df_csv["year"] = df_csv["_time"].dt.strftime("%Y")
        df_csv["month"] = df_csv["_time"].dt.strftime("%m")
        
        # Calculate min and max values per period
        grouped = df_csv.groupby(["year", "month", "_measurement"])
        min_values = grouped["_value"].min().reset_index()
        max_values = grouped["_value"].max().reset_index()
        
        # Calculate consumption per period
        consumption = max_values.copy()
        consumption["period_consumption"] = max_values["_value"] - min_values["_value"]
        
        # Merge with original data
        result = pd.merge(df_csv, consumption[["year", "month", "_measurement", "period_consumption"]],
                         on=["year", "month", "_measurement"])
        
        # Rename columns for output
        result = result.rename(columns={
            "_time": "time",
            "_value": "reading",
            "_measurement": "meter_id"
        })
        
        # Add unit information
        result["reading_unit"] = unit
        result["consumption_unit"] = unit
        
        # Select and order columns
        output_columns = [
            "time", "reading", "reading_unit", 
            "meter_id", "year", "month", 
            "period_consumption", "consumption_unit"
        ]
        
        return result[output_columns].to_csv(index=False)
    except Exception as e:
        if logger:
            logger.error(f"Error generating CSV: {str(e)}")
        raise

def generate_json_output(df, measurement_name, unit, timezone, logger=None):
    """Generate JSON formatted output."""
    try:
        if logger:
            logger.debug("Generating JSON output")
        
        # df['_time'] is already naive Europe/Berlin local time at this point
        df_json = df.copy()

        output_data = {"consumption": []}
        
        # Group by year and month
        df_json["_year"] = df_json["_time"].dt.strftime("%Y")
        df_json["_month"] = df_json["_time"].dt.strftime("%m")
        
        for (year, month), month_df in df_json.groupby(["_year", "_month"]):
            max_val = month_df["_value"].max()
            min_val = month_df["_value"].min()
            consumption = max_val - min_val
            
            meter_values = []
            for _, row in month_df.iterrows():
                meter_values.append({
                    "time": row["_time"].isoformat(),
                    "reading": {
                        "value": round(row["_value"], 4),
                        "unit": unit
                    },
                    "meter_id": row["_measurement"]
                })
            
            output_data["consumption"].append({
                "year": year,
                "month": month,
                "consumption_month": {
                    "value": round(consumption, 4),
                    "unit": unit
                },
                "meter_values": meter_values
            })
        
        return json.dumps(output_data, indent=2)
    except Exception as e:
        if logger:
            logger.error(f"Error generating JSON: {str(e)}")
        raise

def generate_xml_output(df, measurement_name, unit, timezone, logger=None):
    """Generate XML formatted output."""
    try:
        if logger:
            logger.debug("Generating XML output")
        
        # df['_time'] is already naive Europe/Berlin local time at this point
        df_xml = df.copy()

        # Create root element
        root = ET.Element("consumption_data")
        
        # Group by year and month
        df_xml["_year"] = df_xml["_time"].dt.strftime("%Y")
        df_xml["_month"] = df_xml["_time"].dt.strftime("%m")
        
        for (year, month), month_df in df_xml.groupby(["_year", "_month"]):
            max_val = month_df["_value"].max()
            min_val = month_df["_value"].min()
            consumption = max_val - min_val
            
            period = ET.SubElement(root, "period")
            ET.SubElement(period, "year").text = year
            ET.SubElement(period, "month").text = month
            
            consumption_month = ET.SubElement(period, "consumption_month")
            ET.SubElement(consumption_month, "value").text = str(round(consumption, 4))
            ET.SubElement(consumption_month, "unit").text = unit
            
            meter_values = ET.SubElement(period, "meter_values")
            for _, row in month_df.iterrows():
                item = ET.SubElement(meter_values, "item")
                ET.SubElement(item, "time").text = row["_time"].isoformat()
                
                reading = ET.SubElement(item, "reading")
                ET.SubElement(reading, "value").text = str(round(row["_value"], 4))
                ET.SubElement(reading, "unit").text = unit
                
                ET.SubElement(item, "meter_id").text = row["_measurement"]
        
        # Create pretty printed XML
        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")
    except Exception as e:
        if logger:
            logger.error(f"Error generating XML: {str(e)}")
        raise

def generate_output_files(df, output_dir, logger, measurement_name=None, unit="kWh", out_formats=None, timezone=None, add_gaps=False):
    """Generate output files in specified formats."""
    if out_formats is None:
        out_formats = ["excel"]
    
    if "none" in out_formats:
        return []
    
    if measurement_name is None:
        measurement_name = df["_measurement"].str.replace(" ", "_").iloc[0] if "_measurement" in df.columns and not df["_measurement"].empty else "unknown"
    
    output_files = []
    
    if "excel" in out_formats:
        try:
            excel_file = generate_excel_output(df, output_dir, measurement_name, unit, timezone, logger, add_gaps)
            output_files.append(excel_file)
            logger.info(f"Created Excel file: {excel_file}")
        except Exception as e:
            logger.error(f"Failed to generate Excel output: {str(e)}")
    
    if "json" in out_formats:
        try:
            json_file = generate_json_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(json_file)
            logger.info(f"Created JSON file: {json_file}")
        except Exception as e:
            logger.error(f"Failed to generate JSON output: {str(e)}")
    
    if "xml" in out_formats:
        try:
            xml_file = generate_xml_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(xml_file)
            logger.info(f"Created XML file: {xml_file}")
        except Exception as e:
            logger.error(f"Failed to generate XML output: {str(e)}")
    
    if "csv" in out_formats:
        try:
            csv_file = generate_csv_output_to_file(df, output_dir, measurement_name, unit, timezone, logger)
            output_files.append(csv_file)
            logger.info(f"Created CSV file: {csv_file}")
        except Exception as e:
            logger.error(f"Failed to generate CSV output: {str(e)}")
    
    return output_files

def generate_csv_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate CSV output to file."""
    csv_content = generate_csv_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.csv"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(csv_content)
    
    return output_filename

def generate_json_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate JSON output to file."""
    json_content = generate_json_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.json"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(json_content)
    
    return output_filename

def generate_xml_output_to_file(df, output_dir, measurement_name, unit, timezone, logger=None):
    """Generate XML output to file."""
    xml_content = generate_xml_output(df, measurement_name, unit, timezone, logger)
    time_range = f"from_{df['_time'].min().strftime('%Y-%m-%d_%H-%M-%S')}_to_{df['_time'].max().strftime('%Y-%m-%d_%H-%M-%S')}"
    output_filename = f"{measurement_name}_{time_range}.xml"
    
    with open(os.path.join(output_dir, output_filename), 'w') as f:
        f.write(xml_content)
    
    return output_filename

def detect_meter_id(csv_files, logger=None):
    """Read the meter id from a raw export's .json/.xml sibling.

    The per-window export CSVs (id;value;scaler;unit;status;capture_time)
    don't carry the meter id themselves - it only lives once per export in
    the sibling JSON/XML file's top-level "id" (JSON) / id attribute (XML).
    """
    for csv_file in csv_files:
        base = os.path.splitext(csv_file)[0]
        json_file = base + ".json"
        xml_file = base + ".xml"
        if os.path.isfile(json_file):
            with open(json_file) as f:
                data = json.load(f)
            if data.get("id"):
                return data["id"]
        if os.path.isfile(xml_file):
            root_id = ET.parse(xml_file).getroot().attrib.get("id")
            if root_id:
                return root_id
    raise ValueError(
        "Could not determine meter id from any .json/.xml sibling of the "
        "export_*.csv files - pass --meter explicitly"
    )


def get_workbook_cutoff(xlsx_path, logger=None):
    """Return (cutoff, last_sheet_name): the newest reading already in xlsx_path.

    Monthly sheets are named YYYY_MM (sorts correctly as plain strings) and
    store rows newest-first, so the last sheet's row 2 holds the newest
    reading in the whole workbook.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        month_sheets = sorted(s for s in wb.sheetnames if s != "Verbrauch")
        if not month_sheets:
            raise ValueError(f"No monthly sheets found in {xlsx_path}")
        last_sheet = month_sheets[-1]
        ws = wb[last_sheet]
        first_data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if first_data_row is None:
            raise ValueError(f"Sheet '{last_sheet}' in {xlsx_path} has no data rows")
        cutoff = datetime.strptime(first_data_row[0], "%d.%m.%Y %H:%M:%S")
        return cutoff, last_sheet
    finally:
        wb.close()


def load_raw_export_folder(folder, meter, lo_iso, hi_iso, timezone, divisor, logger=None):
    """Normalize a folder of raw export_*.csv files into a processed dataframe.

    Shells out to normalize_meter_csv.awk (the same schema-detecting
    normalizer the rest of the pipeline uses) restricted to [lo_iso, hi_iso),
    then runs the result through the same process_dataframe()/
    convert_to_timezone() steps load_and_process_data() uses for already
    -normalized input.
    """
    csv_files = sorted(glob.glob(os.path.join(folder, "export_*.csv")))
    if not csv_files:
        raise ValueError(f"No export_*.csv files found in {folder}")

    if meter is None:
        meter = detect_meter_id(csv_files, logger)
        if logger:
            logger.info(f"Detected meter id from export siblings: {meter}")

    result = subprocess.run(
        ["awk", "-v", f"lo={lo_iso}", "-v", f"hi={hi_iso}", "-v", f"meter={meter}",
         "-f", NORMALIZE_AWK, "--"] + csv_files,
        capture_output=True, text=True, check=True,
    )

    normalized = "_time;_value;_measurement\n" + result.stdout
    df = pd.read_csv(io.StringIO(normalized), delimiter=';', parse_dates=['_time'])
    if df.empty:
        return df

    df = process_dataframe(df, keep_every=1, timezone=timezone, divisor=divisor, logger=logger)
    df = convert_to_timezone(df, timezone, logger)
    df['_time'] = df['_time'].dt.tz_localize(None)
    return df


MONTH_TAG_RE = re.compile(r'^\d{4}_\d{2}$')
GAP_THRESHOLD = pd.Timedelta(minutes=20)


def find_insertion_row_for_month(ws_consumption, year_month):
    """Row where a new month's Verbrauch row belongs, keeping month rows in
    chronological (YYYY_MM string) order ahead of any trailing hand-added
    content (Summe, blanks, ...) - not just after whatever the last row is,
    since a backfilled month can land earlier than months already present."""
    insert_row = 2
    for r in range(2, ws_consumption.max_row + 1):
        cell_val = str(ws_consumption.cell(row=r, column=1).value or "")
        if not MONTH_TAG_RE.match(cell_val):
            break
        if cell_val < year_month:
            insert_row = r + 1
        else:
            break
    return insert_row


def find_gaps(sorted_desc_times, threshold=GAP_THRESHOLD, timezone='Europe/Berlin'):
    """(gap_start, gap_end, duration) for each consecutive-reading gap over
    `threshold`, given timestamps sorted newest-first.

    A plain naive-datetime diff misreports every DST spring-forward
    transition (e.g. 2025-03-30's clocks jumping 02:00->03:00 CEST) as a
    ~75-minute gap, even though the underlying readings are exactly 15
    minutes apart in real time. Re-localizing to `timezone`-aware and
    diffing that fixes it - but only gets applied to pairs that already
    exceed `threshold` under the naive diff: pytz can't tell which of a
    DST fall-back's two ambiguous occurrences (e.g. 2025-10-26's repeated
    02:00-02:59) a naive time belongs to, so blindly re-localizing every
    pair around a fall-back can manufacture a gap that was never even a
    candidate. Restricting the DST-aware recheck to already-over-threshold
    naive diffs means it can only rule a candidate OUT, never invent one.
    """
    tz = pytz.timezone(timezone)
    gaps = []
    for newer, older in zip(sorted_desc_times, sorted_desc_times[1:]):
        naive_delta = newer - older
        if naive_delta <= threshold:
            continue
        real_delta = tz.localize(newer) - tz.localize(older)
        if real_delta > threshold:
            gaps.append((older, newer, real_delta))
    return gaps


def read_month_rows(ws):
    """(datetime, time_str, value, meas) tuples for a month sheet's data rows."""
    rows = []
    for time_str, value, meas in ws.iter_rows(min_row=2, values_only=True):
        if time_str is None:
            continue
        rows.append((datetime.strptime(time_str, "%d.%m.%Y %H:%M:%S"), time_str, value, meas))
    return rows


def write_month_rows(ws, rows, header_style):
    """Replace a month sheet's data rows (header kept) with `rows`, which
    must already be sorted newest-first."""
    if ws.max_row >= 2:
        ws.delete_rows(2, amount=ws.max_row - 1)
    for _, time_str, value, meas in rows:
        ws.append([time_str, value, meas])
    apply_german_number_format(ws, value_col_idx=2)
    apply_zebra_formatting(ws, header_style)
    width_df = pd.DataFrame(
        [(r[0], r[2], r[3]) for r in rows], columns=['_time', '_value', '_measurement']
    )
    set_dynamic_column_widths(ws, width_df)


def merge_month_sheet(ws, month_df, header_style, timezone='Europe/Berlin'):
    """Merge newly-normalized rows into an existing month sheet.

    Dedupes by exact timestamp (already-present readings are skipped),
    re-sorts newest-first, and rewrites the sheet's data rows. Readings can
    be older than, newer than, or interleaved with what's already there -
    e.g. backfilling an archived day alongside topping up with today's live
    export in the same run.
    """
    existing = read_month_rows(ws)
    known_times = {r[1] for r in existing}
    time_idx = month_df.columns.get_loc('_time')
    value_idx = month_df.columns.get_loc('_value')
    meas_idx = month_df.columns.get_loc('_measurement')

    added = 0
    for row in month_df.itertuples(index=False):
        dt = row[time_idx]
        time_str = dt.strftime("%d.%m.%Y %H:%M:%S")
        if time_str in known_times:
            continue
        existing.append((dt, time_str, row[value_idx], row[meas_idx]))
        known_times.add(time_str)
        added += 1

    existing.sort(key=lambda r: r[0], reverse=True)
    if added:
        write_month_rows(ws, existing, header_style)

    gaps = find_gaps([r[0] for r in existing], timezone=timezone)
    return added, gaps


def find_last_month_row(ws_consumption):
    """Row number of the last Monat row (column A == YYYY_MM), scanning from
    the top rather than assuming it's the sheet's last row - real workbooks
    in this pipeline have hand-added rows (a "Summe" total, blank spacer
    rows, ad-hoc lookup blocks) below the last month."""
    last = 1
    for r in range(2, ws_consumption.max_row + 1):
        if MONTH_TAG_RE.match(str(ws_consumption.cell(row=r, column=1).value or "")):
            last = r
    return last


def extend_summe_formula(ws_consumption, old_last_row, new_last_row, logger=None):
    """Bump a hand-added 'Summe' row's =SUM(B<start>:B<old_last_row>) total
    to cover the newly-inserted month rows too. Only touches cells matching
    that exact pattern - any other custom formula in/after the Summe row is
    left untouched, since this workbook's Verbrauch sheet carries hand-added
    columns/blocks this script has no business rewriting."""
    pattern = re.compile(rf'^=SUM\((\$?[A-Z]+)(\d+):(\$?)B{old_last_row}\)$')
    patched = 0
    for row in ws_consumption.iter_rows():
        for cell in row:
            if cell.value == "Summe" and cell.column == 1:
                for c in row:
                    if isinstance(c.value, str):
                        m = pattern.match(c.value)
                        if m:
                            c.value = f"=SUM({m.group(1)}{m.group(2)}:{m.group(3)}B{new_last_row})"
                            patched += 1
    if logger:
        if patched:
            logger.info(f"Extended {patched} 'Summe' SUM(...) formula(s) to include the new month row(s)")
        else:
            logger.warning(
                "No 'Summe' B-column SUM(...) formula found to extend - if this workbook has a "
                "hand-added total row, double check it still covers the new month row(s)"
            )


def restyle_summe_row(ws_consumption, header_style):
    """Re-apply header-style coloring to a "Summe" row's A/B cells.

    apply_zebra_formatting() has no way to skip specific rows - it
    unconditionally restripes every cell's fill/border in the sheet's used
    range - so this must be called every time that runs, not just when a
    new month was just inserted.
    """
    for row in ws_consumption.iter_rows():
        for cell in row:
            if cell.column == 1 and cell.value == "Summe":
                for col in (1, 2):
                    c = ws_consumption.cell(row=cell.row, column=col)
                    for attr, value in header_style.items():
                        setattr(c, attr, value)


def extend_plausibilitaetstest_formula(ws_consumption, old_last_row, new_last_row, logger=None):
    """Bump add_summe_and_plausibility_rows()'s Plausibilitaetstest row's
    reference to the last month's own sheet (A{old_last_row}) so it still
    points at the newest month after new rows are inserted above it."""
    pattern = re.compile(
        rf'^=INDIRECT\("\'"&A{old_last_row}&"\'!B2"\)-'
        rf'INDEX\(INDIRECT\("\'"&A2&"\'!B:B"\),COUNTA\(INDIRECT\("\'"&A2&"\'!A:A"\)\)\)$'
    )
    patched = 0
    for row in ws_consumption.iter_rows():
        for cell in row:
            if cell.value == PLAUSIBILITAETSTEST_LABEL and cell.column == 1:
                for c in row:
                    if isinstance(c.value, str) and pattern.match(c.value):
                        c.value = (
                            f"""=INDIRECT("'"&A{new_last_row}&"'!B2")-"""
                            f"""INDEX(INDIRECT("'"&A2&"'!B:B"),COUNTA(INDIRECT("'"&A2&"'!A:A")))"""
                        )
                        patched += 1
    if logger:
        if patched:
            logger.info("Extended the Plausibilitaetstest row's reference to the new last month")
        else:
            logger.warning(
                "No Plausibilitaetstest row found to extend - if this workbook predates that "
                "feature, this is expected"
            )


def append_to_workbook(df, xlsx_path, unit, timezone='Europe/Berlin', logger=None):
    """Merge readings into an already-generated workbook, in place.

    This is a real merge, not an append: readings can be older than, newer
    than, or interleaved with what a month's sheet already has (e.g.
    backfilling an archived day alongside topping up with today's live
    export in the same run) - existing months are merged and deduped by
    exact timestamp via merge_month_sheet(). Brand-new months get a fresh
    sheet plus a Verbrauch row inserted in chronological order (not blindly
    at the sheet's end - there may be hand-added rows, e.g. a "Summe" total,
    below the last month). Older months' formulas resolve the next month via
    INDIRECT, so they self-correct once a new month exists.

    Returns (added_rows, gaps): gaps is a list of
    (year_month, gap_start, gap_end, duration) for every >20min gap
    remaining in any month this call touched.
    """
    wb = load_workbook(xlsx_path)
    header_style = build_header_style()
    ws_consumption = wb["Verbrauch"]
    existing_months = {s for s in wb.sheetnames if s != "Verbrauch"}

    df = df.sort_values('_time', ascending=False).copy()
    df["_year_month"] = df["_time"].dt.strftime("%Y_%m")

    last_month_row_before = find_last_month_row(ws_consumption)
    had_trailing_rows_before = last_month_row_before < ws_consumption.max_row
    new_month_count = 0

    added_rows = 0
    all_gaps = []
    for year_month, month_df in df.groupby("_year_month"):
        if year_month in existing_months:
            ws = wb[year_month]
            added, gaps = merge_month_sheet(ws, month_df, header_style, timezone=timezone)
            added_rows += added
            all_gaps.extend((year_month, *g) for g in gaps)
            if logger:
                if added:
                    logger.info(
                        f"Merged {added} new row(s) into existing sheet "
                        f"'{year_month}' (now {ws.max_row - 1} total)"
                    )
                else:
                    logger.info(f"No new rows for existing sheet '{year_month}' (all already present)")
        else:
            insert_row = find_insertion_row_for_month(ws_consumption, year_month)
            ws_consumption.insert_rows(insert_row, amount=1)
            build_month_sheet(wb, ws_consumption, year_month, month_df, unit, header_style, row_num=insert_row)
            existing_months.add(year_month)
            new_month_count += 1
            added_rows += len(month_df)
            all_gaps.extend(
                (year_month, *g) for g in find_gaps(sorted(month_df['_time'], reverse=True), timezone=timezone)
            )
            if logger:
                logger.info(f"Created new sheet '{year_month}' with {len(month_df)} row(s)")

    if new_month_count and had_trailing_rows_before:
        last_month_row_after = find_last_month_row(ws_consumption)
        extend_summe_formula(ws_consumption, last_month_row_before, last_month_row_after, logger)
        extend_plausibilitaetstest_formula(ws_consumption, last_month_row_before, last_month_row_after, logger)
        trailing = ws_consumption.max_row - last_month_row_after
        if trailing and logger:
            logger.warning(
                f"{trailing} row(s) below the last month (blank/Summe/other hand-added content) "
                "were shifted - please manually verify any custom formulas there besides the "
                "Summe total and Plausibilitaetstest row"
            )

    if added_rows:
        apply_zebra_formatting(ws_consumption, header_style)
        # apply_zebra_formatting() blanket-restripes every cell's fill in
        # the sheet's used range - including "Summe"'s header-style A/B
        # coloring - so it has to be reapplied every time this runs, not
        # just when a new month was just added.
        restyle_summe_row(ws_consumption, header_style)
        set_dynamic_column_widths(ws_consumption, df, is_consumption=True)
        wb.save(xlsx_path)
    wb.close()
    return added_rows, all_gaps


def rename_with_new_end(xlsx_path, new_end):
    """Rename ..._to_<old-date>.xlsx to ..._to_<new_end>.xlsx, matching the
    naming convention generate_excel_output() uses for a fresh workbook."""
    dirname, base = os.path.split(xlsx_path)
    new_base = re.sub(
        r'_to_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}\.xlsx$',
        f'_to_{new_end.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx',
        base,
    )
    if new_base == base:
        return xlsx_path
    new_path = os.path.join(dirname, new_base)
    os.rename(xlsx_path, new_path)
    return new_path


def main():
    examples = f"""\
Examples:
  
  # Default behavior (keep all rows)
  %(prog)s --file data.csv
  
  # Basic downsampling scenarios:
  %(prog)s --file data.csv --keep-every 10    # Keep every 10th row
  %(prog)s --file data.csv --keep-every 0     # Only first/last per month
  
  # Mixed output format examples:
  %(prog)s --file data.csv --keep-every 10 --out-format excel csv
  %(prog)s --file data.csv --keep-every 5 --stdout-format json
  
  # Pipeline usage:
  cat data.csv | %(prog)s --stdin --keep-every 20 > output.json
  find ./data -name "*.csv" | xargs %(prog)s --keep-every 30 --out-format excel
"""

    parser = argparse.ArgumentParser(
        description="Process meter data with configurable downsampling",
        epilog=examples,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Input options
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument("--folder", help="Path to folder with CSV files (with --append-to: the raw export_*.csv folder)")
    input_group.add_argument("--file", help="Path to single CSV file")
    input_group.add_argument("--stdin", action="store_true", help="Read from stdin")

    # Append mode: update an existing workbook instead of generating a new one
    parser.add_argument("--append-to", metavar="XLSX",
                       help="Existing workbook to update in place with readings newer than its "
                            "latest sheet, instead of generating a new workbook. Requires --folder "
                            "pointing at a directory of raw export_*.csv/.json/.xml files.")
    parser.add_argument("--meter",
                       help="Meter id to inject for raw exports (with --append-to). "
                            "Default: auto-detected from the export files' .json/.xml siblings.")

    # Processing parameters
    parser.add_argument("--timezone", default="Europe/Berlin", help="Timezone for conversion")
    parser.add_argument("--delimiter", default=",", help="CSV delimiter character")
    parser.add_argument("--pattern", default=r".*\.csv$", help="File pattern regex")
    parser.add_argument("--output", default=".", help="Output directory")
    
    # Column mappings
    parser.add_argument("--time-col", default="_time", help="Time column name")
    parser.add_argument("--value-col", default="_value", help="Value column name")
    parser.add_argument("--measurement-col", default="_measurement", help="Measurement column name")
    
    # Processing options
    parser.add_argument("--divisor", type=float, default=1.0, help="Divide values by this number")
    parser.add_argument("--unit", default="kWh", help="Measurement unit for display")
    
    # Output options
    parser.add_argument("--out-format", nargs="+", default=["excel"],
                       choices=["none", "excel", "csv", "json", "xml"],
                       help="Output file formats")
    parser.add_argument("--add-gaps", action="store_true",
                       help="Add a 'Lücken (keine Daten vom SMGW)' block to the Verbrauch sheet "
                            "listing gaps over 20 minutes in the input data (excel output only; "
                            "not available with --append-to)")
    parser.add_argument("--stdout-format", default="none",
                       choices=["none", "json", "csv", "xml"],
                       help="Output format for stdout")
    
    # Logging
    parser.add_argument("--log-level", default="INFO",
                       choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
                       help="Logging level")

    # Downsampling parameter
    parser.add_argument(
        "--keep-every",
        type=int,
        default=1,
        metavar="N",
        help="""Downsample output by keeping:
1 = all rows (default),
0 = only first/last per month,
N = every Nth row between endpoints
Applies to all output formats""",
    )

    args = parser.parse_args()

    # Validate downsampling parameter
    if args.keep_every < 0:
        parser.error("--keep-every must be >= 0")
    
    logger = setup_logging(getattr(logging, args.log_level))

    if args.append_to:
        if not args.folder:
            parser.error("--append-to requires --folder (the raw export_*.csv directory)")
        try:
            cutoff_before, last_sheet = get_workbook_cutoff(args.append_to, logger)
            logger.info(f"Workbook's latest reading before this run is {cutoff_before} (sheet '{last_sheet}')")

            df = load_raw_export_folder(
                args.folder, args.meter, "0001-01-01T00:00:00Z", "9999-12-31T23:59:59Z",
                args.timezone, args.divisor, logger
            )
            if df.empty:
                logger.info("No data found in folder - workbook left unchanged.")
                return

            added_rows, gaps = append_to_workbook(df, args.append_to, args.unit, timezone=args.timezone, logger=logger)
            if added_rows == 0:
                logger.info("All readings in the folder were already present - workbook left unchanged.")

            if gaps:
                logger.warning(f"{len(gaps)} gap(s) over 20 minutes remain in the merged month(s):")
                for year_month, start, end, duration in sorted(gaps, key=lambda g: g[1]):
                    logger.warning(f"  {year_month}: {start} -> {end} ({duration})")
            else:
                logger.info("No gaps over 20 minutes remain in the merged month(s).")

            if added_rows:
                cutoff_after, _ = get_workbook_cutoff(args.append_to, logger)
                new_path = rename_with_new_end(args.append_to, cutoff_after)
                logger.info(f"Merged {added_rows} new row(s); workbook is now {new_path}")
        except Exception as e:
            logger.error(f"Append failed: {str(e)}", exc_info=logger.level <= logging.DEBUG)
            sys.exit(1)
        return

    try:
        # Load data
        if args.stdin:
            df = load_from_stdin(
                args.time_col, args.value_col, args.measurement_col,
                args.delimiter, args.keep_every, args.timezone, args.divisor, logger
            )
        else:
            column_mapping = {
                '_time': args.time_col,
                '_value': args.value_col,
                '_measurement': args.measurement_col
            }
            df = load_and_process_data(
                args.file or args.folder, args.pattern, args.timezone,
                column_mapping, args.keep_every, args.divisor, logger, args.delimiter
            )
        
        # Handle stdout output
        if args.stdout_format != "none":
            if args.stdout_format == "json":
                try:
                    json_output = generate_json_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(json_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout JSON: {str(e)}")
                    sys.exit(1)
            elif args.stdout_format == "xml":
                try:
                    xml_output = generate_xml_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(xml_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout XML: {str(e)}")
                    sys.exit(1)
            elif args.stdout_format == "csv":
                try:
                    csv_output = generate_csv_output(
                        df, None, args.unit, args.timezone, logger
                    )
                    print(csv_output)
                except Exception as e:
                    logger.error(f"Failed to generate stdout CSV: {str(e)}")
                    sys.exit(1)
        
        # Handle file outputs
        output_files = generate_output_files(
            df, args.output, logger, None, args.unit, args.out_format, args.timezone, args.add_gaps
        )
        
        if not output_files and "none" not in args.out_format:
            logger.error("No output files were created")
            sys.exit(1)
            
    except Exception as e:
        logger.error(f"Processing failed: {str(e)}", exc_info=logger.level <= logging.DEBUG)
        sys.exit(1)

if __name__ == "__main__":
    main()